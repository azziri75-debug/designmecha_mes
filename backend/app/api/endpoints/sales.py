from typing import List, Optional, Union
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import desc, or_
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload, joinedload
from sqlalchemy import delete
from app.api import deps
from app.models.sales import (
    Estimate, EstimateItem, SalesOrder, SalesOrderItem, OrderStatus,
    DeliveryHistory, DeliveryHistoryItem
)
from app.schemas import sales as schemas
from app.models.product import Product, ProductProcess, BOM
from app.models.basics import Partner
from app.models.production import ProductionPlan, ProductionPlanItem, WorkOrder
from app.models.quality import InspectionResult, QualityDefect
from app.models.purchasing import PurchaseOrder, PurchaseStatus, PurchaseOrderItem, OutsourcingOrder, OutsourcingStatus, OutsourcingOrderItem, MaterialRequirement

import uuid
from datetime import datetime, date
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from app.core import config
import json
from app.api.utils.mrp import calculate_and_record_mrp

router = APIRouter()

# --- Estimates ---

@router.post("/estimates/", response_model=schemas.Estimate)
async def create_estimate(
    estimate_in: schemas.EstimateCreate,
    db: AsyncSession = Depends(deps.get_db)
):
    # 1. Create Estimate Header
    db_estimate = Estimate(
        partner_id=estimate_in.partner_id,
        estimate_date=estimate_in.estimate_date or datetime.now().date(),
        valid_until=estimate_in.valid_until,
        total_amount=estimate_in.total_amount,
        note=estimate_in.note,
        attachment_file=estimate_in.attachment_file
    )
    db.add(db_estimate)
    await db.flush() # Get ID

    # 2. Create Items
    try:
        for item in estimate_in.items:
            db_item = EstimateItem(
                estimate_id=db_estimate.id,
                product_id=item.product_id,
                product_name=item.product_name,
                unit_price=item.unit_price,
                quantity=item.quantity,
                note=item.note
            )
            db.add(db_item)
        await db.commit()
    except Exception as e:
        if await deps.ensure_staff_columns(db, e): # Reusing the helper (it now checks SALES_ITEM_COLUMNS too)
            # Retry once
            await db.rollback()
            # We need to re-add everything... this is a bit complex in a single block.
            # But ensure_staff_columns commits the ALTER TABLE.
            # Let's just raise 500 and ask for retry if auto-fix happened.
            raise HTTPException(status_code=500, detail="Database schema was updated. Please try again.")
        raise e
    await db.refresh(db_estimate)
    
    # Eager load for response
    query = select(Estimate).options(
        selectinload(Estimate.items).selectinload(EstimateItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(Estimate.partner)
    ).where(Estimate.id == db_estimate.id)
    result = await db.execute(query)
    return result.scalars().first()

@router.get("/estimates/", response_model=List[schemas.Estimate])
async def read_estimates(
    skip: int = 0,
    limit: int = 100,
    partner_id: Optional[int] = None,
    major_group_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    product_name: Optional[str] = None,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    Retrieve estimates with advanced filtering.
    """
    query = select(Estimate).options(
        selectinload(Estimate.items).selectinload(EstimateItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
        selectinload(Estimate.items).selectinload(EstimateItem.product).selectinload(Product.bom_items).selectinload(BOM.child_product),
        selectinload(Estimate.partner)
    )

    if partner_id:
        query = query.where(Estimate.partner_id == partner_id)
    
    if major_group_id or product_name:
        subquery = select(EstimateItem.estimate_id).join(Product)
        if major_group_id:
            from app.models.product import ProductGroup
            subquery = subquery.join(ProductGroup, Product.group_id == ProductGroup.id)\
                               .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))
        if product_name:
            subquery = subquery.where(
                or_(Product.name.ilike(f"%{product_name}%"), Product.specification.ilike(f"%{product_name}%"))
            )
        query = query.where(Estimate.id.in_(subquery))

    if start_date:
        query = query.where(Estimate.estimate_date >= start_date)
    if end_date:
        query = query.where(Estimate.estimate_date <= end_date)

    query = query.order_by(desc(Estimate.estimate_date)).offset(skip).limit(limit or 2000)

    result = await db.execute(query)
    return result.scalars().all()

@router.get("/estimates/{estimate_id}", response_model=schemas.Estimate)
async def read_estimate(
    estimate_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(Estimate).options(
        selectinload(Estimate.items).selectinload(EstimateItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(Estimate.partner)
    ).where(Estimate.id == estimate_id)
    result = await db.execute(query)
    estimate = result.scalars().first()
    
    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")
    return estimate

@router.post("/estimates/{estimate_id}/export_excel", response_model=schemas.Estimate)
async def export_estimate_excel(
    estimate_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(Estimate).options(
        selectinload(Estimate.items).selectinload(EstimateItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(Estimate.partner)
    ).where(Estimate.id == estimate_id)
    result = await db.execute(query)
    estimate = result.scalars().first()

    if not estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    try:
        from openpyxl.styles import PatternFill
        import urllib.parse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "견적서"

        header_font = Font(name='Malgun Gothic', size=16, bold=True)
        bold_font = Font(name='Malgun Gothic', size=10, bold=True)
        normal_font = Font(name='Malgun Gothic', size=10)
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_align = Alignment(horizontal='right', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        def style_range(ws, cell_range, border=thin_border, font=normal_font, alignment=center_align, fill=None):
            for row in ws[cell_range]:
                for cell in row:
                    cell.border = border
                    cell.font = font
                    cell.alignment = alignment
                    if fill:
                        cell.fill = fill

        col_widths = {'A': 5, 'B': 25, 'C': 15, 'D': 8, 'E': 8, 'F': 12, 'G': 12, 'H': 15}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        ws.merge_cells('A1:H2')
        ws['A1'] = "견  적  서"
        ws['A1'].font = header_font
        ws['A1'].alignment = center_align

        ws.merge_cells('A4:D4')
        ws['A4'] = f"일자: {estimate.estimate_date}"
        ws['A4'].alignment = left_align
        
        ws.merge_cells('A5:D5')
        ws['A5'] = f"수신: {estimate.partner.name if estimate.partner else ''} 귀하"
        ws['A5'].alignment = left_align
        ws['A5'].font = bold_font
        
        ws.merge_cells('A6:D6')
        ws['A6'] = f"합계금액: {estimate.total_amount:,.0f} 원 (VAT 별도)"
        ws['A6'].alignment = left_align
        ws['A6'].font = bold_font
        
        ws.merge_cells('E4:E6')
        ws['E4'] = "공급자"
        ws['E4'].font = bold_font
        ws['E4'].alignment = center_align
        ws['E4'].fill = gray_fill
        ws['E4'].border = thin_border
        
        ws.merge_cells('F4:H4')
        ws['F4'] = "상호: (주)디자인메카"
        ws.merge_cells('F5:H5')
        ws['F5'] = "대표: 조인호"
        ws.merge_cells('F6:H6')
        ws['F6'] = "등록번호: xxx-xx-xxxxx" # placeholder if actual doesn't exist
        
        style_range(ws, 'F4:H6', alignment=left_align)
        
        ws['A8'] = "번호"
        ws['B8'] = "품명"
        ws['C8'] = "규격"
        ws['D8'] = "수량"
        ws['E8'] = "단위"
        ws['F8'] = "단가"
        ws['G8'] = "금액"
        ws['H8'] = "비고"
        style_range(ws, 'A8:H8', font=bold_font, fill=gray_fill)

        row_idx = 9
        for idx, item in enumerate(estimate.items):
            ws.cell(row=row_idx, column=1, value=idx+1)
            
            pname = item.product.name if item.product else "-"
            ws.cell(row=row_idx, column=2, value=pname).alignment = left_align
            
            spec = item.product.specification if item.product else "-"
            ws.cell(row=row_idx, column=3, value=spec)
            
            ws.cell(row=row_idx, column=4, value=item.quantity)
            ws.cell(row=row_idx, column=5, value="EA")
            
            ws.cell(row=row_idx, column=6, value=f"{item.unit_price:,.0f}").alignment = right_align
            ws.cell(row=row_idx, column=7, value=f"{(item.quantity * item.unit_price):,.0f}").alignment = right_align
            
            ws.cell(row=row_idx, column=8, value=item.note or "-").alignment = left_align
            
            style_range(ws, f'A{row_idx}:H{row_idx}')
            row_idx += 1

        ws.merge_cells(f'A{row_idx}:E{row_idx}')
        ws[f'A{row_idx}'] = "합계"
        ws[f'A{row_idx}'].font = bold_font
        ws[f'A{row_idx}'].fill = gray_fill
        ws[f'A{row_idx}'].alignment = center_align
        ws.cell(row=row_idx, column=6, value="")
        ws.cell(row=row_idx, column=7, value=f"{estimate.total_amount:,.0f}").alignment = right_align
        ws.cell(row=row_idx, column=7).font = bold_font
        ws.cell(row=row_idx, column=8, value="")
        style_range(ws, f'A{row_idx}:H{row_idx}')
        
        row_idx += 2
        ws.merge_cells(f'A{row_idx}:A{row_idx+1}')
        ws.merge_cells(f'B{row_idx}:H{row_idx+1}')
        ws[f'A{row_idx}'] = "특기사항"
        ws[f'A{row_idx}'].font = bold_font
        ws[f'A{row_idx}'].fill = gray_fill
        ws[f'A{row_idx}'].alignment = center_align
        ws[f'B{row_idx}'] = estimate.note or ""
        ws[f'B{row_idx}'].alignment = left_align
        style_range(ws, f'A{row_idx}:H{row_idx+1}')

        filename = f"Estimate_{estimate.partner.name if estimate.partner else 'Unk'}_{estimate.estimate_date}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        upload_dir = "uploads/estimates"
        os.makedirs(upload_dir, exist_ok=True)
        file_path = os.path.join(upload_dir, filename)
        
        wb.save(file_path)

        file_url = f"/uploads/estimates/{urllib.parse.quote(filename)}"
        new_file = {"name": filename, "url": file_url}
        
        current_files = []
        if estimate.attachment_file:
            if isinstance(estimate.attachment_file, list):
                current_files = estimate.attachment_file
            elif isinstance(estimate.attachment_file, str):
                try:
                    current_files = json.loads(estimate.attachment_file)
                except:
                    pass
        
        current_files.append(new_file)
        
        estimate.attachment_file = current_files
        
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(estimate, "attachment_file")

        db.add(estimate)
        await db.commit()
        await db.refresh(estimate)
        
        query = select(Estimate).options(
            selectinload(Estimate.items).selectinload(EstimateItem.product).options(
                selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(Product.bom_items).selectinload(BOM.child_product)
            ),
            selectinload(Estimate.partner)
        ).where(Estimate.id == estimate_id)
        result = await db.execute(query)
        return result.scalars().first()

    except Exception as e:
        print(f"Excel Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/estimates/{estimate_id}", response_model=schemas.Estimate)
async def update_estimate(
    estimate_id: int,
    estimate_in: schemas.EstimateUpdate,
    db: AsyncSession = Depends(deps.get_db)
):
    # 1. Fetch existing
    query = select(Estimate).options(
        selectinload(Estimate.items),
        selectinload(Estimate.partner)
    ).where(Estimate.id == estimate_id)
    result = await db.execute(query)
    db_estimate = result.scalars().first()
    
    if not db_estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")

    # 2. Update Header Fields
    update_data = estimate_in.model_dump(exclude_unset=True)
    items_data = update_data.pop("items", None)
    
    for field, value in update_data.items():
        setattr(db_estimate, field, value)

    # 3. Update Items (Full Replace Strategy)
    if items_data is not None:
        # Delete existing items
        # Note: If we had cascade delete on DB level, this might be easier, 
        # but explicit delete is safer for logic control
        for item in db_estimate.items:
            await db.delete(item)
        
        # Add new items
        for item_in in estimate_in.items:
            db_item = EstimateItem(
                estimate_id=db_estimate.id,
                product_id=item_in.product_id,
                product_name=item_in.product_name,
                unit_price=item_in.unit_price,
                quantity=item_in.quantity,
                note=item_in.note
            )
            db.add(db_item)

    await db.commit()
    await db.refresh(db_estimate)
    
    # Re-fetch with all eager loads for response
    query = select(Estimate).options(
        selectinload(Estimate.items).selectinload(EstimateItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(Estimate.partner)
    ).where(Estimate.id == estimate_id)
    result = await db.execute(query)
    return result.scalars().first()

@router.delete("/estimates/{estimate_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_estimate(
    estimate_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(Estimate).options(selectinload(Estimate.items)).where(Estimate.id == estimate_id)
    result = await db.execute(query)
    db_estimate = result.scalars().first()
    
    if not db_estimate:
        raise HTTPException(status_code=404, detail="Estimate not found")
        
    # Delete items first (if not cascading) - SQLAlchemy relationship usually handles this 
    # if cascade="all, delete-orphan" is set. Assuming it might not be, let's rely on DB FK cascade or manual.
    # Checking models... items = relationship("EstimateItem", back_populates="estimate", cascade="all, delete-orphan") is best practice.
    # Use manual delete just in case to be safe without checking model definition deep dive right now.
    for item in db_estimate.items:
        await db.delete(item)
        
    await db.delete(db_estimate)
    await db.commit()
    return None

# --- Orders ---

@router.post("/orders/", response_model=schemas.SalesOrder)
async def create_order(
    order_in: schemas.SalesOrderCreate,
    db: AsyncSession = Depends(deps.get_db)
):
    # Generate Order No
    date_str = datetime.now().strftime("%Y%m%d")
    
    # Robust numbering: get the max order_no for today and increment its sequence
    query = select(SalesOrder.order_no).filter(SalesOrder.order_no.like(f"SO-{date_str}-%")).order_by(desc(SalesOrder.order_no)).limit(1)
    result = await db.execute(query)
    last_order_no = result.scalar()
    
    if last_order_no:
        try:
            last_seq = int(last_order_no.split("-")[-1])
            new_seq = last_seq + 1
        except (ValueError, IndexError):
            new_seq = 1
    else:
        new_seq = 1
        
    order_no = f"SO-{date_str}-{new_seq:03d}"

    db_order = SalesOrder(
        order_no=order_no,
        partner_id=order_in.partner_id,
        order_date=order_in.order_date or datetime.now().date(),
        delivery_date=order_in.delivery_date,
        actual_delivery_date=order_in.actual_delivery_date,
        delivery_method=order_in.delivery_method,
        transaction_date=order_in.transaction_date,
        total_amount=order_in.total_amount,
        note=order_in.note,
        status=order_in.status,
        attachment_file=order_in.attachment_file
    )
    db.add(db_order)
    await db.flush()

    for item in order_in.items:
        db_item = SalesOrderItem(
            order_id=db_order.id,
            product_id=item.product_id,
            product_name=item.product_name,
            unit_price=item.unit_price,
            quantity=item.quantity,
            note=item.note,
            delivered_quantity=item.delivered_quantity
        )
        db.add(db_item)
    
    # MRP 계산 및 부족분 기록
    await calculate_and_record_mrp(db, db_order.id)
    
    await db.commit()
    await db.refresh(db_order)
    
    # Re-fetch with full eager loading for response
    query = select(SalesOrder).options(
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(SalesOrder.partner),
        selectinload(SalesOrder.delivery_histories).selectinload(DeliveryHistory.items)
    ).where(SalesOrder.id == db_order.id)
    result = await db.execute(query)
    
    return result.scalars().first()

@router.get("/orders/", response_model=List[schemas.SalesOrder])
async def read_orders(
    skip: int = 0,
    limit: int = 100,
    partner_id: Optional[int] = None,
    major_group_id: Optional[int] = None,
    status: Optional[str] = None,
    date_type: Optional[str] = "order",
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    product_name: Optional[str] = None,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    Retrieve sales orders with advanced filtering.
    """
    query = select(SalesOrder).options(
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product).selectinload(Product.bom_items).selectinload(BOM.child_product),
        selectinload(SalesOrder.partner),
        selectinload(SalesOrder.delivery_histories).selectinload(DeliveryHistory.items)
    )

    if partner_id:
        query = query.where(SalesOrder.partner_id == partner_id)
    
    if major_group_id or product_name:
        subquery = select(SalesOrderItem.order_id).join(Product)
        if major_group_id:
            from app.models.product import ProductGroup
            subquery = subquery.join(ProductGroup, Product.group_id == ProductGroup.id)\
                               .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))
        if product_name:
            subquery = subquery.where(
                or_(Product.name.ilike(f"%{product_name}%"), Product.specification.ilike(f"%{product_name}%"))
            )
        query = query.where(SalesOrder.id.in_(subquery))

    if status:
        query = query.where(SalesOrder.status == status)

    if start_date:
        if date_type == "delivery":
            query = query.where(SalesOrder.delivery_date >= start_date)
        else:
            query = query.where(SalesOrder.order_date >= start_date)
    if end_date:
        if date_type == "delivery":
            query = query.where(SalesOrder.delivery_date <= end_date)
        else:
            query = query.where(SalesOrder.order_date <= end_date)

    if date_type == "delivery":
        query = query.order_by(desc(SalesOrder.delivery_date)).offset(skip).limit(limit or 2000)
    else:
        query = query.order_by(desc(SalesOrder.order_date)).offset(skip).limit(limit or 2000)
    
    result = await db.execute(query)
    return result.scalars().all()

@router.put("/orders/{order_id}", response_model=schemas.SalesOrder)
async def update_order(
    order_id: int,
    order_in: schemas.SalesOrderUpdate, # Use Update schema
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(SalesOrder).options(
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product)
    ).where(SalesOrder.id == order_id)
    result = await db.execute(query)
    db_order = result.scalars().first()
    
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Update Header
    if order_in.partner_id is not None: db_order.partner_id = order_in.partner_id
    if order_in.order_date is not None: db_order.order_date = order_in.order_date
    if order_in.delivery_date is not None: db_order.delivery_date = order_in.delivery_date
    if order_in.actual_delivery_date is not None: db_order.actual_delivery_date = order_in.actual_delivery_date
    if order_in.delivery_method is not None: db_order.delivery_method = order_in.delivery_method
    if order_in.transaction_date is not None: db_order.transaction_date = order_in.transaction_date
    if order_in.total_amount is not None: db_order.total_amount = order_in.total_amount
    if order_in.note is not None: db_order.note = order_in.note
    if order_in.status is not None: db_order.status = order_in.status
    if order_in.attachment_file is not None: db_order.attachment_file = order_in.attachment_file
    
    # Update Items
    if order_in.items is not None:
        from app.models.production import ProductionPlan, ProductionPlanItem, WorkLogItem
        from sqlalchemy import func
        
        existing_items = {item.id: item for item in db_order.items}
        incoming_items = order_in.items
        kept_item_ids = set()
        
        for item_in in incoming_items:
            if item_in.id and item_in.id in existing_items:
                # 1. Update existing item
                db_item = existing_items[item_in.id]

                # Check if product_id or quantity is being changed
                if db_item.product_id != item_in.product_id or db_item.quantity != item_in.quantity:
                    plan_check = await db.execute(
                        select(func.count(ProductionPlanItem.id))
                        .join(ProductionPlan, ProductionPlanItem.plan_id == ProductionPlan.id)
                        .where(ProductionPlan.order_id == db_order.id)
                        .where(ProductionPlanItem.product_id == db_item.product_id)
                    )
                    if plan_check.scalar() > 0:
                        raise HTTPException(
                            status_code=400,
                            detail=f"이미 진행 중인 생산 계획이 있는 품목의 규격(품목)이나 수량은 직접 수정할 수 없습니다."
                        )

                db_item.product_id = item_in.product_id
                db_item.product_name = item_in.product_name
                db_item.unit_price = item_in.unit_price
                db_item.quantity = item_in.quantity
                db_item.note = item_in.note
                db_item.delivered_quantity = item_in.delivered_quantity or 0
                kept_item_ids.add(db_item.id)
            else:
                # 2. Add new item
                new_db_item = SalesOrderItem(
                    order_id=db_order.id,
                    product_id=item_in.product_id,
                    product_name=item_in.product_name,
                    unit_price=item_in.unit_price,
                    quantity=item_in.quantity,
                    note=item_in.note,
                    delivered_quantity=item_in.delivered_quantity or 0
                )
                db.add(new_db_item)
        
        # 3. Handle deletions for items not in the request
        for item_id, item in existing_items.items():
            if item_id not in kept_item_ids:
                # 🚨 ERP Safeguard: Block deletion if delivery history or work logs exist
                
                # Check Delivery History
                hist_check = await db.execute(
                    select(func.count(DeliveryHistoryItem.id))
                    .where(DeliveryHistoryItem.order_item_id == item_id)
                )
                if hist_check.scalar() > 0:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"품목({item.product.name if item.product else item_id})은(는) 이미 납품 이력이 존재하여 삭제할 수 없습니다. 납품 내역을 먼저 취소해 주세요."
                    )
                
                # Check Production Plan
                plan_check = await db.execute(
                    select(func.count(ProductionPlanItem.id))
                    .join(ProductionPlan, ProductionPlanItem.plan_id == ProductionPlan.id)
                    .where(ProductionPlan.order_id == db_order.id)
                    .where(ProductionPlanItem.product_id == item.product_id)
                )
                if plan_check.scalar() > 0:
                    raise HTTPException(
                        status_code=400, 
                        detail=f"품목({item.product.name if item.product else item_id})은(는) 이미 진행 중인 생산 계획이 존재하여 삭제할 수 없습니다. 관련 생산 계획을 먼저 취소해 주세요."
                    )
                
                await db.delete(item)
        
    await db.commit()
    await db.refresh(db_order)
    
    # Re-fetch with full eager loading (including delivery_histories)
    query = select(SalesOrder).options(
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product).options(
            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(Product.bom_items).selectinload(BOM.child_product)
        ),
        selectinload(SalesOrder.partner),
        selectinload(SalesOrder.delivery_histories).selectinload(DeliveryHistory.items)
    ).where(SalesOrder.id == order_id)
    
    result = await db.execute(query)
    return result.scalars().first()

@router.delete("/orders/{order_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_order(
    order_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(SalesOrder).options(selectinload(SalesOrder.items)).where(SalesOrder.id == order_id)
    result = await db.execute(query)
    db_order = result.scalars().first()
    
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Bug 4 Safety Check: Block if already delivered or status is COMPLETED
    if db_order.status == OrderStatus.DELIVERY_COMPLETED or db_order.actual_delivery_date:
        if db_order.status != OrderStatus.PENDING:
            raise HTTPException(status_code=400, detail="이미 납품이 완료된 수주 건은 삭제할 수 없습니다.")
    
    # Cascade delete 연관 데이터 (의존 관계 역순 및 명시적 삭제)
    # 1. 연관된 생산 계획 ID들 가져오기
    plans_query = select(ProductionPlan.id).where(ProductionPlan.order_id == order_id)
    plans_result = await db.execute(plans_query)
    plan_ids = plans_result.scalars().all()

    # [안전 장치] 이미 발주/외주가 진행된 연관 데이터가 있는지 확인 (PENDING 상태 제외)
    # 1) 직결된 MRP(자재소요량) 기반 발주 확인
    mrp_po_check = await db.execute(
        select(PurchaseOrder.id)
        .join(PurchaseOrderItem, PurchaseOrderItem.purchase_order_id == PurchaseOrder.id)
        .join(MaterialRequirement, MaterialRequirement.id == PurchaseOrderItem.material_requirement_id)
        .where(MaterialRequirement.order_id == order_id)
        .where(PurchaseOrder.status != PurchaseStatus.PENDING)
    )
    if mrp_po_check.scalars().first():
        raise HTTPException(status_code=400, detail="이미 자재 발주가 진행된 수주 건은 삭제할 수 없습니다.")

    if plan_ids:
        # 생산 계획에 엮인 공정 아이템들 식별
        items_query = select(ProductionPlanItem.id).where(ProductionPlanItem.plan_id.in_(plan_ids))
        items_res = await db.execute(items_query)
        item_ids = items_res.scalars().all()

        if item_ids:
            # 2) 생산 계획 공차/부품 발주 확인
            po_check = await db.execute(
                select(PurchaseOrder.id)
                .join(PurchaseOrderItem, PurchaseOrderItem.purchase_order_id == PurchaseOrder.id)
                .where(PurchaseOrderItem.production_plan_item_id.in_(item_ids))
                .where(PurchaseOrder.status != PurchaseStatus.PENDING)
            )
            if po_check.scalars().first():
                raise HTTPException(status_code=400, detail="이미 자재 발주가 진행된 수주 건은 삭제할 수 없습니다.")

            # 3) 외주 발주 확인
            os_check = await db.execute(
                select(OutsourcingOrder.id)
                .join(OutsourcingOrderItem, OutsourcingOrderItem.outsourcing_order_id == OutsourcingOrder.id)
                .where(OutsourcingOrderItem.production_plan_item_id.in_(item_ids))
                .where(OutsourcingOrder.status != OutsourcingStatus.PENDING)
            )
            if os_check.scalars().first():
                raise HTTPException(status_code=400, detail="이미 외주 발주가 진행된 수주 건은 삭제할 수 없습니다.")


    if plan_ids:
        # 2. 연관된 생산 공정(Item) ID들 가져오기
        items_query = select(ProductionPlanItem.id).where(ProductionPlanItem.plan_id.in_(plan_ids))
        items_result = await db.execute(items_query)
        item_ids = items_result.scalars().all()

        if item_ids:
            # 3. 작업 지시 및 검사 결과 삭제
            wo_query = select(WorkOrder.id).where(WorkOrder.plan_item_id.in_(item_ids))
            wo_result = await db.execute(wo_query)
            wo_ids = wo_result.scalars().all()
            
            if wo_ids:
                await db.execute(delete(InspectionResult).where(InspectionResult.work_order_id.in_(wo_ids)))
                await db.execute(delete(WorkOrder).where(WorkOrder.id.in_(wo_ids)))

            # 4. 발주 품목 삭제 (ProductionPlanItem 참조)
            await db.execute(delete(PurchaseOrderItem).where(PurchaseOrderItem.production_plan_item_id.in_(item_ids)))
            await db.execute(delete(OutsourcingOrderItem).where(OutsourcingOrderItem.production_plan_item_id.in_(item_ids)))

            # 5. 불량 내역 삭제 (계획 아이템 기준)
            await db.execute(delete(QualityDefect).where(QualityDefect.plan_item_id.in_(item_ids)))
            
            # 6. 생산 공정 아이템 삭제
            await db.execute(delete(ProductionPlanItem).where(ProductionPlanItem.id.in_(item_ids)))
        
        # 7. 불량 내역 삭제 (계획 기준 - 아이템이 없는 경우 대비)
        await db.execute(delete(QualityDefect).where(QualityDefect.plan_id.in_(plan_ids)))

        # 8. 생산 계획 삭제
        await db.execute(delete(ProductionPlan).where(ProductionPlan.id.in_(plan_ids)))
    
    # 9. 불량 내역 삭제 (수주 기준 직접 링크된 것들)
    await db.execute(delete(QualityDefect).where(QualityDefect.order_id == order_id))
    
    # 10. 수주 품목 삭제
    await db.execute(delete(SalesOrderItem).where(SalesOrderItem.order_id == order_id))
    
    # Bug 4 Fix: Explicitly delete MaterialRequirement for this Order
    await db.execute(delete(MaterialRequirement).where(MaterialRequirement.order_id == order_id))
        
    # 11. 수주 헤더 삭제
    await db.delete(db_order)
    await db.commit()
    return None

# --- History & Helper Endpoints ---

@router.get("/history/price")
async def get_recent_price(
    product_id: int,
    partner_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    Get the most recent unit price for a product and partner.
    Checks SalesOrder first, then Estimate.
    """
    # Check last order
    # Joined eager load or just join? We just need unit_price and date.
    # Joining gives us filtered rows.
    query = select(SalesOrderItem).join(SalesOrder)\
        .options(selectinload(SalesOrderItem.order))\
        .where(SalesOrderItem.product_id == product_id)\
        .where(SalesOrder.partner_id == partner_id)\
        .order_by(desc(SalesOrder.order_date))\
        .limit(1)
        
    result = await db.execute(query)
    last_order_item = result.scalars().first()
    
    if last_order_item:
        return {"price": last_order_item.unit_price, "source": "order", "date": last_order_item.order.order_date}
    
    # Check last estimate
    query = select(EstimateItem).join(Estimate)\
        .options(selectinload(EstimateItem.estimate))\
        .where(EstimateItem.product_id == product_id)\
        .where(Estimate.partner_id == partner_id)\
        .order_by(desc(Estimate.estimate_date))\
        .limit(1)
        
    result = await db.execute(query)
    last_estimate_item = result.scalars().first()
        
    if last_estimate_item:
        return {"price": last_estimate_item.unit_price, "source": "estimate", "date": last_estimate_item.estimate.estimate_date}

    return {"price": 0, "source": None}

@router.get("/history/product/{product_id}")
async def get_product_sales_history(
    product_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    """Get all sales history (Estimates and Orders) for a product"""
    # Estimates
    query_est = select(EstimateItem).join(Estimate)\
        .options(selectinload(EstimateItem.estimate).selectinload(Estimate.partner))\
        .where(EstimateItem.product_id == product_id)\
        .order_by(desc(Estimate.estimate_date))
    result_est = await db.execute(query_est)
    estimates = result_est.scalars().all()

    # Orders
    query_ord = select(SalesOrderItem).join(SalesOrder)\
        .options(selectinload(SalesOrderItem.order).selectinload(SalesOrder.partner))\
        .where(SalesOrderItem.product_id == product_id)\
        .order_by(desc(SalesOrder.order_date))
    result_ord = await db.execute(query_ord)
    orders = result_ord.scalars().all()
    
    history = []
    for e in estimates:
        history.append({
            "type": "ESTIMATE",
            "date": e.estimate.estimate_date,
            "partner_name": e.estimate.partner.name if e.estimate.partner else "Unknown",
            "quantity": e.quantity,
            "unit_price": e.unit_price
        })
    for o in orders:
        history.append({
            "type": "ORDER",
            "date": o.order.order_date,
            "partner_name": o.order.partner.name if o.order.partner else "Unknown",
            "quantity": o.quantity,
            "unit_price": o.unit_price
        })
        
    # Sort by date
    history.sort(key=lambda x: x['date'], reverse=True)
    return history


# --- Delivery History Endpoints ---

@router.post("/orders/{order_id}/delivery", response_model=schemas.DeliveryHistory)
async def create_delivery(
    order_id: int,
    delivery_in: schemas.DeliveryHistoryCreate,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    부분 납품 처리:
    1. SalesOrderItem의 delivered_quantity 업데이트
    2. SalesOrder 상태를 PARTIALLY_DELIVERED 또는 DELIVERED로 변경
    3. DeliveryHistory 및 품목 생성
    """
    order_query = select(SalesOrder).options(selectinload(SalesOrder.items)).where(SalesOrder.id == order_id)
    order_res = await db.execute(order_query)
    db_order = order_res.scalars().first()
    if not db_order:
        raise HTTPException(status_code=404, detail="Order not found")

    # Generate Delivery No
    date_str = datetime.now().strftime("%Y%m%d")
    dh_query = select(DeliveryHistory.delivery_no).where(DeliveryHistory.delivery_no.like(f"DH-{date_str}-%")).order_by(desc(DeliveryHistory.delivery_no)).limit(1)
    dh_res = await db.execute(dh_query)
    last_dh_no = dh_res.scalar()
    new_seq = 1
    if last_dh_no:
        try: new_seq = int(last_dh_no.split("-")[-1]) + 1
        except: pass
    delivery_no = f"DH-{date_str}-{new_seq:03d}"

    # Create Delivery History Header
    db_delivery = DeliveryHistory(
        order_id=order_id,
        delivery_date=delivery_in.delivery_date or datetime.now().date(),
        delivery_no=delivery_no,
        note=delivery_in.note,
        attachment_files=delivery_in.attachment_files,
        statement_json=delivery_in.statement_json,
        supplier_info=delivery_in.supplier_info
    )
    db.add(db_delivery)
    await db.flush()

    # Process Items and Update Order Items
    all_completed = True
    for item_in in delivery_in.items:
        # Find order item
        order_item = next((oi for oi in db_order.items if oi.id == item_in.order_item_id), None)
        if not order_item: continue
        
        # Create History Item
        db_item = DeliveryHistoryItem(
            delivery_id=db_delivery.id,
            order_item_id=item_in.order_item_id,
            quantity=item_in.quantity
        )
        db.add(db_item)
        
        # Update delivered_quantity
        order_item.delivered_quantity += item_in.quantity
        
        # Update Inventory (Deduct stock)
        from app.api.utils.inventory import handle_stock_movement
        from app.models.inventory import TransactionType
        await handle_stock_movement(
            db=db,
            product_id=order_item.product_id,
            quantity=-item_in.quantity,
            transaction_type=TransactionType.OUT,
            reference=delivery_no
        )
        
        # Check if this item is fulfilled
        if order_item.delivered_quantity < order_item.quantity:
            all_completed = False

    # Check remaining items in order
    for oi in db_order.items:
        if oi.delivered_quantity < oi.quantity:
            all_completed = False
            break

    # Update Order Status
    if all_completed:
        db_order.status = OrderStatus.DELIVERY_COMPLETED
        db_order.actual_delivery_date = db_delivery.delivery_date
    else:
        db_order.status = OrderStatus.PARTIALLY_DELIVERED

    # [Fix] 납품 완료 시 연관 생산계획 자동 완료 처리
    try:
        if all_completed:
            plans_res = await db.execute(
                select(ProductionPlan).where(
                    ProductionPlan.order_id == order_id,
                    ProductionPlan.status != 'COMPLETED'
                )
            )
            pending_plans = plans_res.scalars().all()
            if pending_plans:
                pending_plan_ids = [p.id for p in pending_plans]
                for plan in pending_plans:
                    plan.status = 'COMPLETED'
                    db.add(plan)
                ppi_res = await db.execute(
                    select(ProductionPlanItem).where(
                        ProductionPlanItem.plan_id.in_(pending_plan_ids),
                        ProductionPlanItem.status != 'COMPLETED'
                    )
                )
                for ppi in ppi_res.scalars().all():
                    ppi.status = 'COMPLETED'
                    db.add(ppi)
    except Exception as e:
        print(f'[create_delivery] Auto-complete production plan failed: {e}')

    await db.commit()
    await db.refresh(db_delivery)
    
    # Eager load for response
    query = select(DeliveryHistory).options(
        selectinload(DeliveryHistory.items).selectinload(DeliveryHistoryItem.order_item).selectinload(SalesOrderItem.product)
    ).where(DeliveryHistory.id == db_delivery.id)
    res = await db.execute(query)
    return res.scalars().first()

@router.get("/orders/{order_id}/delivery", response_model=List[schemas.DeliveryHistory])
async def get_delivery_histories(
    order_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(DeliveryHistory).options(
        selectinload(DeliveryHistory.items).selectinload(DeliveryHistoryItem.order_item).selectinload(SalesOrderItem.product)
    ).where(DeliveryHistory.order_id == order_id).order_by(desc(DeliveryHistory.delivery_date))
    res = await db.execute(query)
    return res.scalars().all()

@router.put("/orders/{order_id}/delivery/{delivery_id}", response_model=schemas.DeliveryHistory)
async def update_delivery_history(
    order_id: int,
    delivery_id: int,
    delivery_update: schemas.DeliveryHistoryUpdate,
    db: AsyncSession = Depends(deps.get_db)
):
    query = select(DeliveryHistory).where(
        DeliveryHistory.id == delivery_id,
        DeliveryHistory.order_id == order_id
    )
    res = await db.execute(query)
    db_delivery = res.scalars().first()
    if not db_delivery:
        raise HTTPException(status_code=404, detail="Delivery history not found")

    if delivery_update.statement_json is not None:
        db_delivery.statement_json = delivery_update.statement_json
    if delivery_update.supplier_info is not None:
        db_delivery.supplier_info = delivery_update.supplier_info
    
    await db.commit()
    await db.refresh(db_delivery)
    return db_delivery

@router.post("/orders/delivery/{delivery_id}/attach-statement")
async def attach_delivery_statement(
    delivery_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(deps.get_db)
):
    """
    거래명세서 PDF를 납품 이력에 첨부파일로 저장
    """
    query = select(DeliveryHistory).where(DeliveryHistory.id == delivery_id)
    res = await db.execute(query)
    db_delivery = res.scalars().first()
    
    if not db_delivery:
        raise HTTPException(status_code=404, detail="Delivery history not found")

    # Ensure uploads directory exists
    upload_dir = os.path.join("uploads", "statements")
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir, exist_ok=True)

    # Save File
    file_ext = os.path.splitext(file.filename)[1]
    file_name = f"stmt_{delivery_id}_{uuid.uuid4().hex[:8]}{file_ext}"
    file_path = os.path.join(upload_dir, file_name)
    
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)

    # Update attachment_files (JSON list)
    current_files = db_delivery.attachment_files or []
    new_file_url = f"/uploads/statements/{file_name}"
    if new_file_url not in current_files:
        current_files.append(new_file_url)
        db_delivery.attachment_files = current_files
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(db_delivery, "attachment_files")

    await db.commit()
    return {"status": "success", "file_url": new_file_url}

@router.post("/orders/{order_id}/batch-complete")
async def batch_complete_order(
    order_id: int,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    납품 완결 시 후방 공정(생산계획, 발주) 일괄 완료 처리
    """
    # 1. 생산 계획 COMPLETED
    plans_query = select(ProductionPlan).where(ProductionPlan.order_id == order_id)
    plans_res = await db.execute(plans_query)
    plans = plans_res.scalars().all()
    plan_ids = [p.id for p in plans]
    
    for plan in plans:
        plan.status = "COMPLETED"
        db.add(plan)

    # 2. 관련 생산 공정(ProductionPlanItem) COMPLETED
    if plan_ids:
        items_query = select(ProductionPlanItem).where(ProductionPlanItem.plan_id.in_(plan_ids))
        items_res = await db.execute(items_query)
        plan_items = items_res.scalars().all()
        plan_item_ids = [pi.id for pi in plan_items]
        for pi in plan_items:
            pi.status = "COMPLETED"
            db.add(pi)

    # 3. 관련 자재 발주(PurchaseOrder) COMPLETED (PENDING/ORDERED 건)
    # 직접 연결된 PO
    po_query = select(PurchaseOrder).where(
        PurchaseOrder.order_id == order_id,
        PurchaseOrder.status.in_(["PENDING", "ORDERED", "PARTIAL"])
    )
    po_res = await db.execute(po_query)
    for po in po_res.scalars().all():
        po.status = PurchaseStatus.COMPLETED
        db.add(po)

    # 생산계획(PlanItem)을 통해 연결된 PO
    if plan_ids:
        sub_po_query = select(PurchaseOrder).join(PurchaseOrderItem).where(
            PurchaseOrderItem.production_plan_item_id.in_(plan_item_ids),
            PurchaseOrder.status.in_(["PENDING", "ORDERED", "PARTIAL"])
        ).distinct()
        sub_po_res = await db.execute(sub_po_query)
        for po in sub_po_res.scalars().all():
            po.status = PurchaseStatus.COMPLETED
            db.add(po)

    # 4. 관련 외주 발주(OutsourcingOrder) COMPLETED
    if plan_ids:
        os_query = select(OutsourcingOrder).join(OutsourcingOrderItem).where(
            OutsourcingOrderItem.production_plan_item_id.in_(plan_item_ids),
            OutsourcingOrder.status.in_(["PENDING", "ORDERED"])
        ).distinct()
        os_res = await db.execute(os_query)
        for os_card in os_res.scalars().all():
            os_card.status = OutsourcingStatus.COMPLETED
            db.add(os_card)

    # 5. MRP(MaterialRequirement) COMPLETED
    mrp_query = select(MaterialRequirement).where(MaterialRequirement.order_id == order_id)
    mrp_res = await db.execute(mrp_query)
    for mrp in mrp_res.scalars().all():
        mrp.status = "COMPLETED"
        db.add(mrp)

    await db.commit()
    return {"message": "Success", "plan_count": len(plans)}

@router.get("/delivery-status", response_model=List[schemas.DeliveryStatusResponse])
@router.get("/delivery-status/", response_model=List[schemas.DeliveryStatusResponse], include_in_schema=False)
async def read_delivery_status(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    partner_name: Optional[str] = None,
    major_group_id: Optional[int] = None,
    status: Optional[str] = None,
    db: AsyncSession = Depends(deps.get_db)
):
    """
    납품 현황 조회를 위한 수주 목록 (배송 이력 포함)
    """
    query = select(SalesOrder).options(
        # partner (required by SalesOrder schema)
        selectinload(SalesOrder.partner),
        # items -> product (required by SalesOrderItem schema)
        selectinload(SalesOrder.items).selectinload(SalesOrderItem.product),
        # delivery_histories -> items -> order_item -> product
        selectinload(SalesOrder.delivery_histories).selectinload(
            DeliveryHistory.items
        ).selectinload(
            DeliveryHistoryItem.order_item
        ).selectinload(SalesOrderItem.product),
    ).order_by(desc(SalesOrder.order_date))

    if start_date:
        query = query.where(SalesOrder.order_date >= start_date)
    if end_date:
        query = query.where(SalesOrder.order_date <= end_date)
    if partner_name:
        query = query.join(Partner).where(Partner.name.ilike(f"%{partner_name}%"))
    if major_group_id:
        from app.models.product import ProductGroup
        query = query.join(SalesOrderItem).join(Product).join(ProductGroup, Product.group_id == ProductGroup.id)\
                     .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))\
                     .distinct()
    if status and status != 'ALL':
        query = query.where(SalesOrder.status == status)

    result = await db.execute(query)
    return result.scalars().unique().all()


# ─────────────────────────────────────────────────────────────
# 거래명세서 PDF 첨부 (납품 이력 → statement_json 업데이트)
# ─────────────────────────────────────────────────────────────
@router.post("/delivery-histories/{history_id}/attach-statement")
async def attach_statement_to_delivery_history(
    history_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(deps.get_db),
    current_user=Depends(deps.get_current_user),
):
    """거래명세서 PDF를 저장하고 해당 DeliveryHistory에 링크를 저장합니다."""
    # 1. Fetch delivery history
    result = await db.execute(select(DeliveryHistory).where(DeliveryHistory.id == history_id))
    history = result.scalars().first()
    if not history:
        raise HTTPException(status_code=404, detail="DeliveryHistory not found")

    # 2. Save PDF file to disk
    _BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    save_dir = os.path.join(_BASE_DIR, "uploads", "statements")
    os.makedirs(save_dir, exist_ok=True)

    filename = f"statement_{history_id}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf"
    file_path = os.path.join(save_dir, filename)
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    # 3. Build public URL
    file_url = f"/static/statements/{filename}"

    # 4. Update statement_json on the history record
    existing = history.statement_json or {}
    if isinstance(existing, str):
        import json as _json
        try:
            existing = _json.loads(existing)
        except Exception:
            existing = {}
    existing["pdf_url"] = file_url
    existing["attached_at"] = datetime.now().isoformat()
    history.statement_json = existing

    await db.commit()
    await db.refresh(history)

    return {"status": "ok", "pdf_url": file_url, "history_id": history_id}


# ─────────────────────────────────────────────────────────────
# 납품 이력 수정 (PUT)
# ─────────────────────────────────────────────────────────────
@router.put("/delivery-histories/{history_id}", response_model=schemas.DeliveryHistory)
async def update_delivery_history(
    history_id: int,
    delivery_update: schemas.DeliveryHistoryUpdate,
    db: AsyncSession = Depends(deps.get_db),
    current_user=Depends(deps.get_current_user),
):
    """납품 이력 수정 API (수량 변경 시 수주 잔량 및 상태 동기화 포함)"""
    result = await db.execute(
        select(DeliveryHistory)
        .options(
            selectinload(DeliveryHistory.items).selectinload(DeliveryHistoryItem.order_item),
            selectinload(DeliveryHistory.order).selectinload(SalesOrder.items)
        )
        .where(DeliveryHistory.id == history_id)
    )
    history = result.scalars().first()
    if not history:
        raise HTTPException(status_code=404, detail="DeliveryHistory not found")

    # 1) 기본 정보 수정
    if delivery_update.note is not None:
        history.note = delivery_update.note
    if delivery_update.delivery_date is not None:
        history.delivery_date = delivery_update.delivery_date
    if delivery_update.statement_json is not None:
        history.statement_json = delivery_update.statement_json
    if delivery_update.supplier_info is not None:
        history.supplier_info = delivery_update.supplier_info

    # 2) 품목 수량 수정 및 수주 데이터 동기화
    if delivery_update.items is not None:
        for item_in in delivery_update.items:
            # 기존 납품 내역 아이템 찾기
            dh_item = next((it for it in history.items if it.order_item_id == item_in.order_item_id), None)
            if not dh_item:
                continue # 혹은 새로 추가하는 로직이 필요할 수 있으나 여기선 수정만 처리
            
            old_qty = dh_item.quantity
            new_qty = item_in.quantity
            diff = new_qty - old_qty
            
            if diff != 0:
                # 납품 이력 아이템 수량 업데이트
                dh_item.quantity = new_qty
                # 수주 품목의 누적 납품 수량 업데이트
                if dh_item.order_item:
                    dh_item.order_item.delivered_quantity = max(0, (dh_item.order_item.delivered_quantity or 0) + diff)
                    
                    # Update Inventory (Deduct stock by diff)
                    from app.api.utils.inventory import handle_stock_movement
                    from app.models.inventory import TransactionType
                    await handle_stock_movement(
                        db=db,
                        product_id=dh_item.order_item.product_id,
                        quantity=-diff,
                        transaction_type=TransactionType.OUT if diff > 0 else TransactionType.ADJUSTMENT,
                        reference=history.delivery_no
                    )

        # 3) 수주 상태 재계산
        order = history.order
        if order:
            total_qty = sum(it.quantity for it in order.items)
            delivered_qty = sum(it.delivered_quantity or 0 for it in order.items)
            
            if delivered_qty >= total_qty:
                order.status = OrderStatus.DELIVERY_COMPLETED
                order.actual_delivery_date = history.delivery_date
            elif delivered_qty > 0:
                order.status = OrderStatus.PARTIALLY_DELIVERED
            else:
                order.status = OrderStatus.CONFIRMED

    await db.commit()
    await db.refresh(history)
    
    # Re-fetch with options after refresh to avoid MissingGreenlet
    result = await db.execute(
        select(DeliveryHistory)
        .options(
            selectinload(DeliveryHistory.items).selectinload(DeliveryHistoryItem.order_item).selectinload(SalesOrderItem.product),
            selectinload(DeliveryHistory.order).selectinload(SalesOrder.items).selectinload(SalesOrderItem.product),
            selectinload(DeliveryHistory.order).selectinload(SalesOrder.partner)
        )
        .where(DeliveryHistory.id == history_id)
    )
    return result.scalars().first()


# ─────────────────────────────────────────────────────────────
# 납품 이력 삭제/취소 (DELETE) - 수주 잔량 복원 포함
# ─────────────────────────────────────────────────────────────
@router.delete("/delivery-histories/{history_id}")
async def delete_delivery_history(
    history_id: int,
    db: AsyncSession = Depends(deps.get_db),
    current_user=Depends(deps.get_current_user),
):
    """납품 이력 삭제 API (삭제 시 수주 잔량 원복 포함)"""
    result = await db.execute(
        select(DeliveryHistory)
        .options(selectinload(DeliveryHistory.items).selectinload(DeliveryHistoryItem.order_item))
        .where(DeliveryHistory.id == history_id)
    )
    history = result.scalars().first()
    if not history:
        raise HTTPException(status_code=404, detail="DeliveryHistory not found")

    order_id = history.order_id

    # 1) 납품된 수량을 수주 품목의 delivered_quantity에서 차감 (원복)
    for dh_item in history.items:
        order_item = dh_item.order_item
        if order_item:
            order_item.delivered_quantity = max(0, (order_item.delivered_quantity or 0) - dh_item.quantity)
            
            # Update Inventory (Revert stock: Add back)
            from app.api.utils.inventory import handle_stock_movement
            from app.models.inventory import TransactionType
            await handle_stock_movement(
                db=db,
                product_id=order_item.product_id,
                quantity=dh_item.quantity,
                transaction_type=TransactionType.IN,
                reference=f"Delete DH ({history.delivery_no})"
            )

    # 2) 납품 이력 삭제
    await db.delete(history)
    await db.flush()

    # 3) 해당 수주의 상태 재계산
    order_result = await db.execute(
        select(SalesOrder)
        .options(selectinload(SalesOrder.items))
        .where(SalesOrder.id == order_id)
    )
    order = order_result.scalars().first()
    if order:
        remaining_histories = await db.execute(
            select(DeliveryHistory).where(DeliveryHistory.order_id == order_id)
        )
        has_deliveries = remaining_histories.scalars().first() is not None

        if not has_deliveries:
            order.status = OrderStatus.CONFIRMED
        else:
            total_qty = sum(it.quantity for it in order.items)
            delivered_qty = sum(it.delivered_quantity or 0 for it in order.items)
            if delivered_qty >= total_qty:
                order.status = OrderStatus.DELIVERY_COMPLETED
            else:
                order.status = OrderStatus.PARTIALLY_DELIVERED

    await db.commit()
    return {"status": "ok", "message": f"DeliveryHistory {history_id} deleted and quantities reverted"}

# --- TEMPORARY FIX ENDPOINT ---
@router.get("/fix-mismatched-plans")
async def fix_mismatched_plans_in_db(db: AsyncSession = Depends(deps.get_db)):
    """
    TEMPORARY ENDPOINT to auto-fix data corruption.
    """
    from sqlalchemy import text
    try:
        query = text("""
        UPDATE production_plan_items
        SET 
            product_id = (
                SELECT soi.product_id 
                FROM sales_orders so 
                JOIN sales_order_items soi ON so.id = soi.order_id 
                JOIN production_plans pp ON so.id = pp.order_id 
                WHERE pp.id = production_plan_items.plan_id
                LIMIT 1
            ),
            quantity = (
                SELECT soi.quantity 
                FROM sales_orders so 
                JOIN sales_order_items soi ON so.id = soi.order_id 
                JOIN production_plans pp ON so.id = pp.order_id 
                WHERE pp.id = production_plan_items.plan_id
                LIMIT 1
            )
        WHERE id IN (
            SELECT ppi.id
            FROM sales_orders so
            JOIN sales_order_items soi ON so.id = soi.order_id
            JOIN production_plans pp ON so.id = pp.order_id
            JOIN production_plan_items ppi ON pp.id = ppi.plan_id
            WHERE soi.product_id != ppi.product_id OR soi.quantity != ppi.quantity
        );
        """)
        await db.execute(query)
        await db.commit()
        return {"status": "success", "message": "어긋난 생산계획 품목 및 수량 데이터가 수주 데이터에 맞춰 정상적으로 동기화되었습니다. 새로고침해 주세요."}
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
