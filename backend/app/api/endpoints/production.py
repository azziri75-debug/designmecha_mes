from typing import Any, List, Optional, Union
from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import select, cast, String, text, func, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.exc import MultipleResultsFound

from app.api import deps
from app.core.timezone import now_kst
from app.models.production import ProductionPlan, ProductionPlanItem, ProductionStatus, WorkLog, WorkLogItem
from app.models.sales import SalesOrder, SalesOrderItem, OrderStatus
from app.models.product import Product, ProductProcess, Process, BOM, ProductGroup
from app.models.purchasing import (
    PurchaseOrderItem, OutsourcingOrderItem, PurchaseOrder, OutsourcingOrder, 
    PurchaseStatus, OutsourcingStatus, MaterialRequirement
)
from app.models.basics import Partner, Staff, Equipment
from app.models.inventory import StockProduction, Stock, StockProductionStatus, TransactionType
from app.api.utils.inventory import handle_stock_movement, handle_backflush
from app.api.utils.status_cascade import on_production_item_completed
from app.schemas import production as schemas
from datetime import datetime, date
import uuid
import json
import os
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter()


async def sync_plan_item_status(db: AsyncSession, plan_item_id: int):
    """
    Update ProductionPlanItem status based on WorkLogItem quantities.
    """
    result = await db.execute(
        select(WorkLogItem)
        .where(WorkLogItem.plan_item_id == plan_item_id)
    )
    items = result.scalars().all()
    
    total_good = sum(item.good_quantity for item in items)
    
    plan_item = await db.get(ProductionPlanItem, plan_item_id)
    if not plan_item:
        return

    if total_good >= plan_item.quantity:
        # [NEW] 상태 연계 유틸리티 호출 (발주 완료 및 재고 반영 포함)
        await on_production_item_completed(db, plan_item, reference=f"WorkLog (PI#{plan_item_id})")
    elif total_good > 0:
        plan_item.status = ProductionStatus.IN_PROGRESS
        db.add(plan_item)
    elif total_good == 0 and plan_item.status in [ProductionStatus.IN_PROGRESS, ProductionStatus.COMPLETED]:
        plan_item.status = ProductionStatus.PLANNED
        db.add(plan_item)
    
    await db.flush()
    
    # --- Auto-Complete Check for Parent Plan ---
    await check_and_complete_production_plan(db, plan_item.plan_id)

async def check_and_complete_production_plan(db: AsyncSession, plan_id: int):
    """
    모든 공정이 완료되었는지 확인하고, 그렇다면 생산 계획을 완료 처리합니다.
    """
    result = await db.execute(
        select(ProductionPlan).options(selectinload(ProductionPlan.items)).where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    if not plan or not plan.items:
        return

    all_completed = all(item.status == ProductionStatus.COMPLETED for item in plan.items)
    if all_completed and plan.status != ProductionStatus.COMPLETED:
        await update_production_plan_status(plan_id, ProductionStatus.COMPLETED, db)

async def process_stock_deduction(db: AsyncSession, plan_id: int):
    """
    생산 계획 확정 시, 사용하기로 한 재고(stock_use_quantity)를 실제 창고에서 차감합니다.
    공정별로 중복 차감되지 않도록 품목별로 합산하여 한 번만 차감합니다.
    """
    result = await db.execute(
        select(ProductionPlan).options(selectinload(ProductionPlan.items)).where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    if not plan:
        return

    # 1. 품목별 차감 수량 집계 (중복 방지)
    deductions = {} # {product_id: stock_use_quantity}
    items_to_mark = []
    
    for item in plan.items:
        if (item.stock_use_quantity or 0) > 0 and not item.stock_deducted:
            # 모든 공정에서 동일한 stock_use_quantity를 가지므로 덮어쓰기 방식으로 집계 가능
            deductions[item.product_id] = item.stock_use_quantity
            items_to_mark.append(item)

    if not deductions:
        return

    from app.api.utils.inventory import handle_stock_movement
    from app.models.inventory import TransactionType

    # 2. 실제 재고 차감 실행
    for pid, qty in deductions.items():
        await handle_stock_movement(
            db=db,
            product_id=pid,
            quantity=-qty,
            transaction_type=TransactionType.OUT,
            reference=f"생산계획 확정 재고소진 (Plan #{plan.id})"
        )

    # 3. 모든 관련 품목에 대해 차감 완료 표시
    for item in items_to_mark:
        item.stock_deducted = True
        db.add(item)
    
    await db.flush()

async def _handle_production_completion_effects(db: AsyncSession, plan: ProductionPlan):
    """
    생산 계획 완료 시 발생하는 부수 효과(수주 동기화, 재고 입고, 백플러시 등)를 처리합니다.
    """
    from app.api.utils.inventory import handle_stock_movement, handle_backflush
    from app.models.inventory import TransactionType
    from app.models.purchasing import PurchaseOrder, PurchaseStatus, OutsourcingOrder, OutsourcingStatus
    from app.models.sales import OrderStatus

    # 1. 하위 구매/외주 품목도 자동으로 '완료' 처리 (만약 미완료 상태라면)
    affected_po_ids = set()
    affected_oo_ids = set()
    for item in plan.items:
        # 1. Update Purchase Items
        for po_item in item.purchase_items:
            if po_item.quantity > po_item.received_quantity:
                po_item.received_quantity = po_item.quantity
                db.add(po_item)
            affected_po_ids.add(po_item.purchase_order_id)
            
        # 2. Update Outsourcing Items
        for oo_item in item.outsourcing_items:
            if oo_item.status != OutsourcingStatus.COMPLETED:
                oo_item.status = OutsourcingStatus.COMPLETED
                db.add(oo_item)
            affected_oo_ids.add(oo_item.outsourcing_order_id)
    
    await db.flush()
    
    # 2. Check and Complete Purchase Orders
    for po_id in affected_po_ids:
        po_query = select(PurchaseOrder).options(selectinload(PurchaseOrder.items)).where(PurchaseOrder.id == po_id)
        po_result = await db.execute(po_query)
        po = po_result.scalars().first()
        if po:
            all_received = all(i.received_quantity >= i.quantity for i in po.items)
            if all_received:
                po.status = PurchaseStatus.COMPLETED
                po.delivery_date = now_kst().date()
                db.add(po)

    # 3. Check and Complete Outsourcing Orders
    for oo_id in affected_oo_ids:
        oo_query = select(OutsourcingOrder).options(selectinload(OutsourcingOrder.items)).where(OutsourcingOrder.id == oo_id)
        oo_result = await db.execute(oo_query)
        oo = oo_result.scalars().first()
        if oo:
            all_completed = all(i.status == OutsourcingStatus.COMPLETED for i in oo.items)
            if all_completed:
                oo.status = OutsourcingStatus.COMPLETED
                oo.delivery_date = now_kst().date()
                db.add(oo)

    # 4. Aggregate Net Production Quantities from Plan Items
    # [FIX] Use max() to handle potential inconsistency across processes of the same product
    net_quantities = {} # {product_id: net_qty}
    for item in plan.items:
        current_max = net_quantities.get(item.product_id, 0)
        net_quantities[item.product_id] = max(current_max, item.quantity or 0)

    # 5. Stock Movement & Backflush Hook
    if plan.stock_production:
        sp = plan.stock_production
        prod_qty = net_quantities.get(sp.product_id, sp.quantity)
        
        await handle_stock_movement(
            db=db, product_id=sp.product_id, quantity=prod_qty,
            transaction_type=TransactionType.IN, reference=sp.production_no
        )
        await handle_backflush(
            db=db, parent_product_id=sp.product_id, produced_quantity=prod_qty, reference=sp.production_no
        )

        stock_query = select(Stock).where(Stock.product_id == sp.product_id)
        s_res = await db.execute(stock_query)
        stock = s_res.scalars().first()
        if stock and stock.in_production_quantity >= prod_qty:
            stock.in_production_quantity -= prod_qty
        
        sp.status = StockProductionStatus.COMPLETED
        db.add(sp)
        
    elif plan.order:
        for pid, qty in net_quantities.items():
            await handle_stock_movement(
                db=db, product_id=pid, quantity=qty,
                transaction_type=TransactionType.IN, reference=plan.order.order_no
            )
            await handle_backflush(
                db=db, parent_product_id=pid, produced_quantity=qty, reference=plan.order.order_no
            )

            stock_query = select(Stock).where(Stock.product_id == pid)
            s_res = await db.execute(stock_query)
            stock = s_res.scalars().first()
            if stock and stock.in_production_quantity >= qty:
                stock.in_production_quantity -= qty
        
        if plan.order.status not in [OrderStatus.DELIVERY_COMPLETED, OrderStatus.PARTIALLY_DELIVERED, "DELIVERED"]:
            plan.order.status = OrderStatus.PRODUCTION_COMPLETED
            db.add(plan.order)
    
    await db.flush()

async def sync_plan_item_cost(db: AsyncSession, plan_item: ProductionPlanItem):
    """
    If INTERNAL, fetch standard cost from ProductProcess and sync plan_item.cost.
    Uses case-insensitive and trimmed matching for process names for robustness.
    """
    if plan_item.course_type == "INTERNAL" and plan_item.product_id and plan_item.process_name:
        # Find corresponding ProductProcess
        stmt = (
            select(ProductProcess.cost)
            .join(Process, ProductProcess.process_id == Process.id)
            .where(
                ProductProcess.product_id == plan_item.product_id,
                func.lower(func.trim(Process.name)) == func.lower(func.trim(plan_item.process_name))
            )
        )
        result = await db.execute(stmt)
        std_unit_cost = result.scalars().first()
        if std_unit_cost is not None:
            plan_item.cost = std_unit_cost * (plan_item.quantity or 1)
            db.add(plan_item)
            await db.flush()

@router.get("/plans", response_model=List[schemas.ProductionPlan])
async def read_production_plans(
    skip: int = 0,
    limit: int = 1000,
    worker_id: Optional[int] = None,
    status: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    partner_id: Optional[int] = None,
    product_name: Optional[str] = None,
    customer_id: Optional[int] = None,
    major_group_id: Optional[int] = None,
    db: AsyncSession = Depends(deps.get_db),
):
    """
    Retrieve production plans with advanced filtering.
    """
    stmt = select(ProductionPlan).outerjoin(SalesOrder).outerjoin(StockProduction)
    
    if worker_id:
        subquery = select(ProductionPlanItem.plan_id).where(ProductionPlanItem.worker_id == worker_id)
        stmt = stmt.where(ProductionPlan.id.in_(subquery))
    
    if status:
        stmt = stmt.where(ProductionPlan.status == status)
    if start_date:
        stmt = stmt.where(ProductionPlan.plan_date >= start_date)
    if end_date:
        stmt = stmt.where(ProductionPlan.plan_date <= end_date)
    if partner_id:
        stmt = stmt.where(or_(
            SalesOrder.partner_id == partner_id,
            StockProduction.partner_id == partner_id
        ))
    if customer_id:
        stmt = stmt.where(SalesOrder.partner_id == customer_id)
    if product_name:
        subquery = select(ProductionPlanItem.plan_id).join(Product).where(Product.name.ilike(f"%{product_name}%"))
        stmt = stmt.where(ProductionPlan.id.in_(subquery))
    if major_group_id:
        subquery = select(ProductionPlanItem.plan_id).join(Product).join(ProductGroup, Product.group_id == ProductGroup.id)\
                   .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))
        stmt = stmt.where(ProductionPlan.id.in_(subquery))

    result = await db.execute(
        stmt
        .options(
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product).options(
                selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(Product.bom_items).selectinload(BOM.child_product)
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.equipment),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.worker),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
            selectinload(ProductionPlan.order).options(
                selectinload(SalesOrder.partner),
                selectinload(SalesOrder.items).selectinload(SalesOrderItem.product)
            ),
            selectinload(ProductionPlan.stock_production).options(
                selectinload(StockProduction.product),
                selectinload(StockProduction.partner)
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.plan).options(
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).options(
                    selectinload(StockProduction.product),
                    selectinload(StockProduction.partner)
                )
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.work_log_items).options(
                selectinload(WorkLogItem.work_log).selectinload(WorkLog.worker),
                selectinload(WorkLogItem.worker)
            )
        )
        .offset(skip)
        .limit(limit)
    )
    plans = result.scalars().all()
    
    # Calculate completed_quantity for each item
    for plan in plans:
        for item in plan.items:
            item.completed_quantity = calculate_completed_quantity(item)
            
    return plans

@router.post("/plans", response_model=schemas.ProductionPlan)
async def create_production_plan(
    plan_in: schemas.ProductionPlanCreate,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Create a production plan from a Sales Order or Stock Production.
    Auto-generates plan items based on Product Processes.
    """
    try:
        # 1. Check if Order or StockProduction exists
        if plan_in.order_id:
            result = await db.execute(select(SalesOrder).where(SalesOrder.id == plan_in.order_id))
            order = result.scalars().first()
            if not order:
                raise HTTPException(status_code=404, detail="Sales Order not found")
                
            # Check if ACTIVE Plan already exists for this order
            result = await db.execute(select(ProductionPlan).where(
                ProductionPlan.order_id == plan_in.order_id,
                cast(ProductionPlan.status, String) != ProductionStatus.CANCELED.value
            ).limit(1))
            if result.scalars().first():
                # Return existing active plan instead of error (Idempotency)
                result = await db.execute(
                    select(ProductionPlan)
                    .options(
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product).options(
                            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                            selectinload(Product.bom_items).selectinload(BOM.child_product)
                        ),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.equipment),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.worker),
                        selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                        selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order)
                    )
                    .where(ProductionPlan.order_id == plan_in.order_id, cast(ProductionPlan.status, String) != ProductionStatus.CANCELED.value)
                    .limit(1)
                )
                return result.scalars().first()
        elif plan_in.stock_production_id:
            result = await db.execute(select(StockProduction).where(StockProduction.id == plan_in.stock_production_id))
            sp = result.scalars().first()
            if not sp:
                raise HTTPException(status_code=404, detail="Stock Production request not found")
            
            # Check if ACTIVE Plan already exists for this stock production
            result = await db.execute(select(ProductionPlan).where(
                ProductionPlan.stock_production_id == plan_in.stock_production_id,
                cast(ProductionPlan.status, String) != ProductionStatus.CANCELED.value
            ).limit(1))
            if result.scalars().first():
                # Return existing active plan instead of error (Idempotency)
                result = await db.execute(
                    select(ProductionPlan)
                    .options(
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product).options(
                            selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                            selectinload(Product.bom_items).selectinload(BOM.child_product)
                        ),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.equipment),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.worker),
                        selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                        selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                        selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order)
                    )
                    .where(ProductionPlan.stock_production_id == plan_in.stock_production_id, cast(ProductionPlan.status, String) != ProductionStatus.CANCELED.value)
                    .limit(1)
                )
                return result.scalars().first()
        else:
            raise HTTPException(status_code=400, detail="Either order_id or stock_production_id is required")

        # 3. Create Plan Header
        plan = ProductionPlan(
            order_id=plan_in.order_id,
            stock_production_id=plan_in.stock_production_id,
            plan_date=plan_in.plan_date,
            status=ProductionStatus.IN_PROGRESS
        )
        db.add(plan)
        await db.flush()
        
        # 4. Generate Items
        if plan_in.items:
            for item_in in plan_in.items:
                plan_item = ProductionPlanItem(
                    plan_id=plan.id,
                    product_id=item_in.product_id,
                    process_name=item_in.process_name,
                    sequence=item_in.sequence,
                    course_type=item_in.course_type,
                    partner_name=item_in.partner_name,
                    work_center=item_in.work_center,
                    estimated_time=item_in.estimated_time,
                    start_date=item_in.start_date,
                    end_date=item_in.end_date,
                    worker_id=item_in.worker_id,
                    equipment_id=item_in.equipment_id,
                    note=item_in.note,
                    status=item_in.status,
                    attachment_file=item_in.attachment_file,
                    quantity=item_in.quantity,
                    gross_quantity=item_in.gross_quantity or item_in.quantity,
                    stock_use_quantity=item_in.stock_use_quantity or 0,
                    cost=item_in.cost
                )
                db.add(plan_item)
                await sync_plan_item_cost(db, plan_item)
            
            await db.flush()
            
            # [FIX] Check for Auto-Completion (Fully satisfied by stock)
            is_fully_stock_satisfied = all((item.quantity or 0) == 0 for item in plan_in.items) if plan_in.items else False
            if is_fully_stock_satisfied:
                plan.status = ProductionStatus.COMPLETED
                # [FIX] 하위 모든 공정도 함께 완료 처리 (발주 대기 목록 노출 방지)
                for itm in plan.items:
                    itm.status = ProductionStatus.COMPLETED
            elif plan.status == ProductionStatus.IN_PROGRESS: 
                plan.status = ProductionStatus.CONFIRMED

            await db.commit()

            # 5. Trigger side effects based on status
            if plan.status == ProductionStatus.COMPLETED:
                await process_stock_deduction(db, plan_id=plan.id)
                # Re-fetch with needed relationships for effects
                res = await db.execute(select(ProductionPlan).options(
                    selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items),
                    selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items),
                    selectinload(ProductionPlan.order),
                    selectinload(ProductionPlan.stock_production)
                ).where(ProductionPlan.id == plan.id))
                refetched_plan = res.scalars().first()
                await _handle_production_completion_effects(db, refetched_plan)
                await db.commit()
            elif plan.status == ProductionStatus.CONFIRMED:
                await process_stock_deduction(db, plan_id=plan.id)
                from app.api.utils.mrp import calculate_and_record_mrp
                await calculate_and_record_mrp(db, plan_id=plan.id)
        else:
            # Default logic for Sales Order (already exists) or Stock Production
            # If stock production, it's usually just one product.
            if plan_in.stock_production_id:
                # Fetch StockProduction to get product_id and quantity
                res = await db.execute(select(StockProduction).where(StockProduction.id == plan_in.stock_production_id))
                sp = res.scalars().first()
                
                # Fetch processes and Product
                stmt = (
                    select(ProductProcess, Process.name, Process.course_type, Product.drawing_file)
                    .join(Process, ProductProcess.process_id == Process.id)
                    .join(Product, ProductProcess.product_id == Product.id)
                    .where(ProductProcess.product_id == sp.product_id)
                    .order_by(ProductProcess.sequence)
                )
                result = await db.execute(stmt)
                processes = result.all()
                
                import json
                for proc, proc_name, proc_course_type, prod_drawing in processes:
                    # Merge attachments: process attachment + product drawing
                    final_attachments = []
                    
                    # Add product drawing files
                    if prod_drawing:
                        try:
                            parsed = json.loads(prod_drawing) if isinstance(prod_drawing, str) else prod_drawing
                            if isinstance(parsed, list): final_attachments.extend(parsed)
                            else: final_attachments.append(parsed)
                        except: final_attachments.append(prod_drawing)
                    
                    # Add process attachment files
                    if proc.attachment_file:
                        try:
                            parsed = json.loads(proc.attachment_file) if isinstance(proc.attachment_file, str) else proc.attachment_file
                            if isinstance(parsed, list): final_attachments.extend(parsed)
                            else: final_attachments.append(parsed)
                        except: final_attachments.append(proc.attachment_file)
                    
                    # Deduplicate and format
                    unique_attachments = []
                    seen_urls = set()
                    for att in final_attachments:
                        if isinstance(att, dict) and att.get('url'):
                            if att['url'] not in seen_urls:
                                unique_attachments.append(att)
                                seen_urls.add(att['url'])
                        elif isinstance(att, str):
                            if att not in seen_urls:
                                unique_attachments.append(att)
                                seen_urls.add(att)

                    final_attachment_json = json.dumps(unique_attachments, ensure_ascii=False) if unique_attachments else None

                    final_course_type = proc.course_type or proc_course_type or "INTERNAL"
                    plan_item = ProductionPlanItem(
                        plan_id=plan.id,
                        product_id=sp.product_id,
                        process_name=proc_name,
                        sequence=proc.sequence,
                        course_type=final_course_type,
                        partner_name=proc.partner_name,
                        work_center=proc.equipment_name,
                        estimated_time=proc.estimated_time,
                        attachment_file=final_attachment_json,
                        quantity=sp.quantity,
                        status=ProductionStatus.PLANNED,
                        cost=(getattr(proc, 'cost', 0) or 0) * sp.quantity
                    )
                    db.add(plan_item)
            else:
                # Sales Order logic (restored)
                result = await db.execute(select(SalesOrderItem).where(SalesOrderItem.order_id == plan_in.order_id))
                order_items = result.scalars().all()
                for item in order_items:
                    stmt = (
                        select(ProductProcess, Process.name, Process.course_type, Product.drawing_file)
                        .join(Process, ProductProcess.process_id == Process.id)
                        .join(Product, ProductProcess.product_id == Product.id)
                        .where(ProductProcess.product_id == item.product_id)
                        .order_by(ProductProcess.sequence)
                    )
                    result = await db.execute(stmt)
                    processes = result.all()
                    import json
                    for proc, proc_name, proc_course_type, prod_drawing in processes:
                        # Merge attachments: process attachment + product drawing
                        final_attachments = []
                        
                        # Add product drawing files
                        if prod_drawing:
                            try:
                                parsed = json.loads(prod_drawing) if isinstance(prod_drawing, str) else prod_drawing
                                if isinstance(parsed, list): final_attachments.extend(parsed)
                                else: final_attachments.append(parsed)
                            except: final_attachments.append(prod_drawing)
                        
                        # Add process attachment files
                        if proc.attachment_file:
                            try:
                                parsed = json.loads(proc.attachment_file) if isinstance(proc.attachment_file, str) else proc.attachment_file
                                if isinstance(parsed, list): final_attachments.extend(parsed)
                                else: final_attachments.append(parsed)
                            except: final_attachments.append(proc.attachment_file)
                        
                        # Deduplicate
                        unique_attachments = []
                        seen_urls = set()
                        for att in final_attachments:
                            if isinstance(att, dict) and att.get('url'):
                                if att['url'] not in seen_urls:
                                    unique_attachments.append(att)
                                    seen_urls.add(att['url'])
                            elif isinstance(att, str):
                                if att not in seen_urls:
                                    unique_attachments.append(att)
                                    seen_urls.add(att)

                        final_attachment_json = json.dumps(unique_attachments, ensure_ascii=False) if unique_attachments else None

                        final_course_type = proc.course_type or proc_course_type or "INTERNAL"
                        plan_item = ProductionPlanItem(
                            plan_id=plan.id,
                            product_id=item.product_id,
                            process_name=proc_name,
                            sequence=proc.sequence,
                            course_type=final_course_type,
                            partner_name=proc.partner_name,
                            work_center=proc.equipment_name,
                            estimated_time=proc.estimated_time,
                            attachment_file=final_attachment_json,
                            quantity=item.quantity,
                            status=ProductionStatus.PLANNED,
                            cost=(getattr(proc, 'cost', 0) or 0) * item.quantity
                        )
                        db.add(plan_item)

        await db.commit()
        await db.refresh(plan)
        
        # Reload logic (Update options to include stock_production)
        result = await db.execute(
            select(ProductionPlan)
            .options(
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product).options(
                    selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                    selectinload(Product.bom_items).selectinload(BOM.child_product)
                ),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.equipment),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.worker),
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).options(
                    selectinload(StockProduction.product),
                    selectinload(StockProduction.partner)
                ),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.plan).options(
                    selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
                ),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.work_log_items).options(
                    selectinload(WorkLogItem.work_log).selectinload(WorkLog.worker),
                    selectinload(WorkLogItem.worker)
                )
            )
            .where(ProductionPlan.id == plan.id)
        )
        plan = result.scalars().first()
        for item in plan.items:
            item.completed_quantity = calculate_completed_quantity(item)
        return plan
    except MultipleResultsFound:
        raise HTTPException(status_code=400, detail="데이터 중복 오류: 동일 품목에 대해 여러 개의 수주 또는 마스터 데이터가 발견되었습니다. 관리자에게 문의하세요.")

def calculate_completed_quantity(item: ProductionPlanItem) -> int:
    """
    Calculate completed quantity based on course type and status.
    For Purchase/Outsourcing, COMPLETED status means 100% progress.
    Otherwise, sum from work logs.
    """
    if item.course_type in ["PURCHASE", "OUTSOURCING"] and item.status == ProductionStatus.COMPLETED:
        return item.quantity
    
    # Check if work_log_items is loaded to avoid MissingGreenlet in async
    from sqlalchemy.orm import attributes
    state = attributes.instance_state(item)
    if "work_log_items" in state.unloaded:
        return 0
        
    try:
        return sum(wl.good_quantity for wl in item.work_log_items)
    except:
        return 0

@router.put("/plans/{plan_id}", response_model=schemas.ProductionPlan)
async def update_production_plan(
    plan_id: int,
    plan_in: schemas.ProductionPlanUpdate,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Update a production plan. Preserves existing items to maintain links to orders.
    """
    # 1. Fetch Plan with items
    result = await db.execute(
        select(ProductionPlan)
        .options(selectinload(ProductionPlan.items))
        .where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Production Plan not found")

    # 2. Update Header & Items
    update_data = plan_in.model_dump(exclude_unset=True)
    
    # [보강] 이미 COMPLETED인 경우 '계획 확정' 등으로 인한 상태 역행 방지
    if plan.status == ProductionStatus.COMPLETED and update_data.get("status") == ProductionStatus.CONFIRMED:
        update_data["status"] = ProductionStatus.COMPLETED
        print(f"Update: Prevented status reversion from COMPLETED to CONFIRMED for Plan #{plan_id}")

    items_data = update_data.pop("items", None)
    
    for field, value in update_data.items():
        setattr(plan, field, value)

    # --- [NEW] 만약 생산 계획 헤더가 COMPLETED로 변경되었다면 하위 모든 공정도 완료 처리 및 날짜 기록 ---
    if update_data.get("status") == ProductionStatus.COMPLETED:
        plan.actual_completion_date = now_kst().date()
        for item in plan.items:
            if item.status != ProductionStatus.COMPLETED:
                # from app.api.endpoints.production import on_production_item_completed # Same file
                await on_production_item_completed(db, item, reference=f"Manual Plan Completion (Plan #{plan_id})")

    # --- [NEW] 공정 상세 항목(items) 업데이트 로직 ---
    if plan_in.items is not None:
        current_items_map = {item.id: item for item in plan.items}
        
        # 1. 삭제 및 업데이트 처리
        # incoming_ids = {item_in.id for item_in in plan_in.items if item_in.id} # Pydantic objects use dot
        incoming_ids = {item_in.id for item_in in plan_in.items if item_in.id}
        for item_id, item in list(current_items_map.items()):
            if item_id not in incoming_ids:
                await db.delete(item)
        
        # 2. 업데이트 및 추가 처리
        for item_in in plan_in.items:
            if item_in.id and item_in.id in current_items_map:
                # 기존 항목 업데이트
                existing_item = current_items_map[item_in.id]
                item_update_dict = item_in.model_dump(exclude_unset=True, exclude={"id"})
                for k, v in item_update_dict.items():
                    setattr(existing_item, k, v)
                await sync_plan_item_cost(db, existing_item)
            else:
                # 신규 항목 추가
                new_item = ProductionPlanItem(
                    plan_id=plan_id,
                    **item_in.model_dump(exclude={"id"})
                )
                db.add(new_item)
                await sync_plan_item_cost(db, new_item)

    await db.flush()
    await db.commit()
    
    # --- Trigger Side Effects ---
    # [FIX] Check for Auto-Completion if items were updated
    if plan_in.items:
        all_qty_zero = all((i.quantity or 0) == 0 for i in plan_in.items)
        if all_qty_zero:
            plan.status = ProductionStatus.COMPLETED


    if plan.status == ProductionStatus.COMPLETED:
        await process_stock_deduction(db, plan_id=plan.id)
        # Re-fetch with needed relationships for effects
        res = await db.execute(select(ProductionPlan).options(
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items),
            selectinload(ProductionPlan.order),
            selectinload(ProductionPlan.stock_production)
        ).where(ProductionPlan.id == plan.id))
        refetched_plan = res.scalars().first()
        await _handle_production_completion_effects(db, refetched_plan)
        await db.commit()
    elif plan.status == ProductionStatus.CONFIRMED:

        await process_stock_deduction(db, plan_id=plan.id)
        from app.api.utils.mrp import calculate_and_record_mrp
        await calculate_and_record_mrp(db, plan_id=plan.id)
    
    # 4. Re-fetch with full options for response
    result = await db.execute(
        select(ProductionPlan)
        .options(
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product).options(
                selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(Product.bom_items).selectinload(BOM.child_product)
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.equipment),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.worker),
            selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
            selectinload(ProductionPlan.stock_production).options(
                selectinload(StockProduction.product),
                selectinload(StockProduction.partner)
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.plan).options(
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
            ),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.work_log_items).options(
                selectinload(WorkLogItem.work_log).selectinload(WorkLog.worker),
                selectinload(WorkLogItem.worker)
            )
        )
        .where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    for item in plan.items:
        item.completed_quantity = calculate_completed_quantity(item)
    return plan

@router.post("/plans/{plan_id}/export_excel", response_model=schemas.ProductionPlan)
async def export_production_plan_excel(
    plan_id: int,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Generate an Excel file for the Production Plan and attach it.
    """
    # 1. Fetch Plan with all relations
    result = await db.execute(
        select(ProductionPlan)
        .options(
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.product),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
            selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
            selectinload(ProductionPlan.order).selectinload(SalesOrder.partner)
        )
        .where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Production Plan not found")

    # 2. Parse Metadata
    metadata = {}
    if plan.sheet_metadata:
        try:
            metadata = json.loads(plan.sheet_metadata) if isinstance(plan.sheet_metadata, str) else plan.sheet_metadata
        except:
            pass
            
    order = plan.order

    # 3. Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "생산관리시트"

    # Define Styles
    header_font = Font(name='Malgun Gothic', size=14, bold=True)
    bold_font = Font(name='Malgun Gothic', size=10, bold=True)
    normal_font = Font(name='Malgun Gothic', size=10)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    gray_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    def style_range(ws, cell_range, border=thin_border, font=normal_font, alignment=center_align, fill=None):
        for row in ws[cell_range]:
            for cell in row:
                cell.border = border
                cell.font = font
                cell.alignment = alignment
                if fill:
                    cell.fill = fill

    col_widths = {'A': 10, 'B': 10, 'C': 15, 'D': 25, 'E': 15, 'F': 12, 'G': 12, 'H': 10, 'I': 10, 'J': 10}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells('A1:J2')
    title_cell = ws['A1']
    title_cell.value = "생산관리시트"
    title_cell.font = header_font
    title_cell.alignment = center_align

    ws.merge_cells('B4:E4')
    ws.merge_cells('G4:J4')
    ws['A4'] = "고객"
    ws['B4'] = order.partner.name if order and order.partner else "-"
    ws['F4'] = "수주일"
    ws['G4'] = str(order.order_date) if order and order.order_date else "-"
    
    ws.merge_cells('B5:E5')
    ws.merge_cells('G5:J5')
    ws['A5'] = "품명"
    unique_products = []
    seen = set()
    for item in plan.items:
        if item.product and item.product.id not in seen:
            seen.add(item.product.id)
            unique_products.append({"product": item.product, "quantity": item.quantity, "note": item.note})
            
    summary_prod_name = "-"
    if unique_products:
        summary_prod_name = unique_products[0]["product"].name
        if len(unique_products) > 1:
            summary_prod_name += f" 외 {len(unique_products) - 1}건"
            
    ws['B5'] = summary_prod_name
    ws['F5'] = "요구납기일"
    ws['G5'] = str(order.delivery_date) if order and order.delivery_date else "-"

    ws.merge_cells('B6:E6')
    ws.merge_cells('G6:J6')
    ws['A6'] = "수주금액"
    ws['B6'] = metadata.get('order_amount', str(order.total_amount) if order else "-")
    ws['F6'] = "수주담당자"
    ws['G6'] = metadata.get('manager', "-")

    style_range(ws, 'A4:A6', font=bold_font, fill=gray_fill)
    style_range(ws, 'F4:F6', font=bold_font, fill=gray_fill)
    style_range(ws, 'A4:J6')

    start_row = 8
    ws.merge_cells(f'A{start_row}:C{start_row}')
    ws.merge_cells(f'D{start_row}:F{start_row}')
    ws.merge_cells(f'G{start_row}:H{start_row}')
    ws.merge_cells(f'I{start_row}:J{start_row}')
    
    ws[f'A{start_row}'] = "품명"
    ws[f'D{start_row}'] = "규격"
    ws[f'G{start_row}'] = "재질"
    ws[f'I{start_row}'] = "수량"
    style_range(ws, f'A{start_row}:J{start_row}', font=bold_font, fill=gray_fill)

    curr_row = start_row + 1
    for prod_info in unique_products:
        ws.merge_cells(f'A{curr_row}:C{curr_row}')
        ws.merge_cells(f'D{curr_row}:F{curr_row}')
        ws.merge_cells(f'G{curr_row}:H{curr_row}')
        ws.merge_cells(f'I{curr_row}:J{curr_row}')
        
        ws[f'A{curr_row}'] = prod_info["product"].name
        ws[f'D{curr_row}'] = prod_info["product"].specification or "-"
        ws[f'G{curr_row}'] = prod_info["product"].material or "-"
        ws[f'I{curr_row}'] = prod_info["quantity"]
        
        style_range(ws, f'A{curr_row}:J{curr_row}', alignment=center_align)
        ws[f'A{curr_row}'].alignment = left_align
        curr_row += 1

    while curr_row < start_row + 4:
        ws.merge_cells(f'A{curr_row}:C{curr_row}')
        ws.merge_cells(f'D{curr_row}:F{curr_row}')
        ws.merge_cells(f'G{curr_row}:H{curr_row}')
        ws.merge_cells(f'I{curr_row}:J{curr_row}')
        style_range(ws, f'A{curr_row}:J{curr_row}')
        curr_row += 1

    memo_row = curr_row + 1
    ws.merge_cells(f'A{memo_row}:A{memo_row+1}')
    ws.merge_cells(f'B{memo_row}:J{memo_row+1}')
    ws[f'A{memo_row}'] = "Memo"
    ws[f'B{memo_row}'] = metadata.get('memo', "")
    style_range(ws, f'A{memo_row}:A{memo_row+1}', font=bold_font, fill=gray_fill)
    style_range(ws, f'B{memo_row}:J{memo_row+1}', alignment=left_align)

    proc_start_row = memo_row + 3
    headers = ["구분", "순번", "공정", "공정내용", "업체", "품명", "규격", "수량", "시작", "종료"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=proc_start_row, column=i+1, value=h)
        cell.font = bold_font
        cell.fill = gray_fill
        cell.border = thin_border
        cell.alignment = center_align

    proc_row = proc_start_row + 1
    def get_type_label(ctype):
        if not ctype: return "-"
        if "INTERNAL" in ctype or "자가" in ctype: return "자가"
        if "OUTSOURCING" in ctype or "외주" in ctype: return "외주"
        if "PURCHASE" in ctype or "구매" in ctype: return "구매"
        return ctype
        
    for idx, item in enumerate(plan.items):
        ws.cell(row=proc_row, column=1, value=get_type_label(item.course_type))
        ws.cell(row=proc_row, column=2, value=item.sequence or (idx + 1))
        ws.cell(row=proc_row, column=3, value=item.process_name or "-")
        
        note_cell = ws.cell(row=proc_row, column=4, value=item.note or "-")
        note_cell.alignment = left_align
        
        partner_cell = ws.cell(row=proc_row, column=5, value=item.partner_name or "-")
        partner_cell.alignment = left_align
        
        ws.cell(row=proc_row, column=6, value="-")
        ws.cell(row=proc_row, column=7, value="-")
        ws.cell(row=proc_row, column=8, value=item.quantity)
        ws.cell(row=proc_row, column=9, value=str(item.start_date) if item.start_date else "-")
        ws.cell(row=proc_row, column=10, value=str(item.end_date) if item.end_date else "-")
        
        style_range(ws, f'A{proc_row}:J{proc_row}')
        proc_row += 1

    upload_dir = "uploads/production"
    os.makedirs(upload_dir, exist_ok=True)
    
    filename = f"ProductionSheet_{plan.id}_{now_kst().strftime('%Y%m%d_%H%M%S')}.xlsx"
    file_path = os.path.join(upload_dir, filename)
    wb.save(file_path)

    file_url = f"/uploads/production/{urllib.parse.quote(filename)}"
    new_attachment = {
        "url": file_url,
        "name": filename
    }

    current_attachments = []
    if plan.attachment_file:
        try:
           current_attachments = json.loads(plan.attachment_file) if isinstance(plan.attachment_file, str) else plan.attachment_file
           if not isinstance(current_attachments, list):
               current_attachments = [current_attachments]
        except:
            pass

    current_attachments.append(new_attachment)
    plan.attachment_file = current_attachments

    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(plan, "attachment_file")

    db.add(plan)
    await db.commit()
    
    result = await db.execute(
        select(ProductionPlan)
        .options(
            selectinload(ProductionPlan.items).options(
                selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(ProductionPlanItem.equipment),
                selectinload(ProductionPlanItem.worker),
                selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
                selectinload(ProductionPlanItem.work_log_items).options(
                    selectinload(WorkLogItem.work_log).selectinload(WorkLog.worker),
                    selectinload(WorkLogItem.worker)
                )
            ),
            selectinload(ProductionPlan.order).selectinload(SalesOrder.partner)
        )
        .where(ProductionPlan.id == plan.id)
    )
    return result.scalars().first()


@router.delete("/plans/{plan_id}", status_code=200)
async def delete_production_plan(
    plan_id: int,
    db: AsyncSession = Depends(deps.get_db),
):
    """
    Delete a production plan and related data (MRP, linked Purchase/Outsourcing Items).
    Blocks if any linked material is already COMPLETED.
    """
    # 1. Fetch Plan with all relevant info for safeguard check
    result = await db.execute(
        select(ProductionPlan)
        .options(
            selectinload(ProductionPlan.order).selectinload(SalesOrder.items),
            selectinload(ProductionPlan.stock_production)
        )
        .where(ProductionPlan.id == plan_id)
    )
    plan = result.scalars().first()
    
    if not plan:
        raise HTTPException(status_code=404, detail="Production Plan not found")

    from sqlalchemy.orm import joinedload
    from app.models.quality import QualityDefect
    from app.models.production import WorkOrder, ProductionStatus
    from app.models.inventory import Stock, TransactionType
    from app.api.utils.inventory import handle_stock_movement, handle_backflush

    # 1-1. 재고 롤백 처리 (삭제 전 실행)
    if plan.status == ProductionStatus.COMPLETED:
        if plan.stock_production:
            sp = plan.stock_production
            await handle_stock_movement(db, sp.product_id, -sp.quantity, TransactionType.OUT, f"Delete Rollback ({sp.production_no})")
            await handle_backflush(db, sp.product_id, -sp.quantity, f"Delete Rollback ({sp.production_no})")
        elif plan.order:
            for item in plan.order.items:
                await handle_stock_movement(db, item.product_id, -item.quantity, TransactionType.OUT, f"Delete Rollback ({plan.order.order_no})")
                await handle_backflush(db, item.product_id, -item.quantity, f"Delete Rollback ({plan.order.order_no})")

    elif plan.status in [ProductionStatus.IN_PROGRESS, ProductionStatus.CONFIRMED]:
        # 생산 중 수량 차감
        if plan.stock_production:
            sp = plan.stock_production
            stock = (await db.execute(select(Stock).where(Stock.product_id == sp.product_id))).scalars().first()
            if stock: stock.in_production_quantity = max(0, stock.in_production_quantity - sp.quantity)
        elif plan.order:
            for item in plan.order.items:
                stock = (await db.execute(select(Stock).where(Stock.product_id == item.product_id))).scalars().first()
                if stock: stock.in_production_quantity = max(0, stock.in_production_quantity - item.quantity)

    # 2. Collect all linked records to check status and handle deletion
    # MRPs linked to the plan
    mrp_stmt = select(MaterialRequirement).where(MaterialRequirement.plan_id == plan_id)
    mrp_res = await db.execute(mrp_stmt)
    mrps = mrp_res.scalars().all()
    mrp_ids = [m.id for m in mrps]
    
    # Plan Items
    pi_stmt = select(ProductionPlanItem).where(ProductionPlanItem.plan_id == plan_id)
    pi_res = await db.execute(pi_stmt)
    plan_items = pi_res.scalars().all()
    plan_item_ids = [pi.id for pi in plan_items]

    all_po_items = []
    all_oo_items = []

    # Gather Purchase Items via MRP
    if mrp_ids:
        po_via_mrp = await db.execute(
            select(PurchaseOrderItem).options(joinedload(PurchaseOrderItem.purchase_order))
            .where(PurchaseOrderItem.material_requirement_id.in_(mrp_ids))
        )
        all_po_items.extend(po_via_mrp.scalars().all())

    # Gather Purchase/Outsourcing Items via Plan Items
    if plan_item_ids:
        po_via_pi = await db.execute(
            select(PurchaseOrderItem).options(joinedload(PurchaseOrderItem.purchase_order))
            .where(PurchaseOrderItem.production_plan_item_id.in_(plan_item_ids))
        )
        all_po_items.extend(po_via_pi.scalars().all())
        
        oo_via_pi = await db.execute(
            select(OutsourcingOrderItem).options(joinedload(OutsourcingOrderItem.outsourcing_order))
            .where(OutsourcingOrderItem.production_plan_item_id.in_(plan_item_ids))
        )
        all_oo_items.extend(oo_via_pi.scalars().all())

    # 🚨 ERP Safeguard: Check for COMPLETED status
    for p_item in all_po_items:
        if p_item.purchase_order and p_item.purchase_order.status == PurchaseStatus.COMPLETED:
            raise HTTPException(
                status_code=400, 
                detail="이미 입고 완료된 자재가 있어 생산계획을 삭제할 수 없습니다. 입고를 먼저 취소해 주세요."
            )
    
    for o_item in all_oo_items:
        if o_item.outsourcing_order and o_item.outsourcing_order.status == OutsourcingStatus.COMPLETED:
            raise HTTPException(
                status_code=400, 
                detail="이미 입고 완료된 외주 건이 있어 생산계획을 삭제할 수 없습니다. 입고를 먼저 취소해 주세요."
            )

    # 3. Cascading Deletion
    affected_po_ids = set(p.purchase_order_id for p in all_po_items if p.purchase_order_id)
    affected_oo_ids = set(o.outsourcing_order_id for o in all_oo_items if o.outsourcing_order_id)

    # Delete Purchase/Outsourcing Items
    # Note: Use a set of unique item IDs to avoid double-deletion if linked to BOTH plan and MRP
    unique_po_ids = set()
    for p_item in all_po_items:
        if p_item.id not in unique_po_ids:
            await db.delete(p_item)
            unique_po_ids.add(p_item.id)
            
    unique_oo_ids = set()
    for o_item in all_oo_items:
        if o_item.id not in unique_oo_ids:
            await db.delete(o_item)
            unique_oo_ids.add(o_item.id)
    
    await db.flush()

    # Delete empty Order headers (headers with no items left)
    from sqlalchemy import func
    for po_id in affected_po_ids:
        rem_res = await db.execute(select(func.count(PurchaseOrderItem.id)).where(PurchaseOrderItem.purchase_order_id == po_id))
        if rem_res.scalar() == 0:
            po = await db.get(PurchaseOrder, po_id)
            if po: await db.delete(po)

    for oo_id in affected_oo_ids:
        rem_res = await db.execute(select(func.count(OutsourcingOrderItem.id)).where(OutsourcingOrderItem.outsourcing_order_id == oo_id))
        if rem_res.scalar() == 0:
            oo = await db.get(OutsourcingOrder, oo_id)
            if oo: await db.delete(oo)

    # Delete MRPs
    for m in mrps:
        await db.delete(m)

    # 4. Cleanup other plan-linked entities
    # Delete Quality Defects
    qd_stmt = select(QualityDefect).where(QualityDefect.plan_id == plan_id)
    qd_res = await db.execute(qd_stmt)
    for qd in qd_res.scalars().all():
        await db.delete(qd)
        
    # Delete Work Orders
    if plan_item_ids:
        wo_stmt = select(WorkOrder).where(WorkOrder.plan_item_id.in_(plan_item_ids))
        wo_res = await db.execute(wo_stmt)
        for wo in wo_res.scalars().all():
            await db.delete(wo)

    # Update linked sales order status (rollback to CONFIRMED)
    if plan.order:
        plan.order.status = OrderStatus.CONFIRMED
        db.add(plan.order)

    # Finally, delete the plan itself
    await db.delete(plan)
    await db.commit()
    
    return {"message": "Production Plan and all related data (MRP, unreceived orders) deleted successfully"}


@router.patch("/plans/{plan_id}/status", response_model=schemas.ProductionPlan)
async def update_production_plan_status(
    plan_id: int,
    status: ProductionStatus,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Update production plan status.
    If COMPLETED, update SalesOrder status to PRODUCTION_COMPLETED.
    """
    try:
        # Eager load items and linked orders
        result = await db.execute(
            select(ProductionPlan)
            .options(
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product),
                selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.purchase_items).options(
                    selectinload(PurchaseOrderItem.purchase_order).selectinload(PurchaseOrder.items)
                ),
                selectinload(ProductionPlan.items).selectinload(ProductionPlanItem.outsourcing_items).options(
                    selectinload(OutsourcingOrderItem.outsourcing_order).selectinload(OutsourcingOrder.items)
                ),
                selectinload(ProductionPlan.order).selectinload(SalesOrder.items).selectinload(SalesOrderItem.product)
            )
            .where(ProductionPlan.id == plan_id)
        )
        plan = result.scalars().first()
        
        if not plan:
            raise HTTPException(status_code=404, detail="Production Plan not found")
            
        old_status = plan.status
        plan.status = status
        
        # 0. CONFIRMED Trigger: Calculate MRP directly inline
        if status == ProductionStatus.CONFIRMED and old_status != ProductionStatus.CONFIRMED:
            print(f"MRP Generation Started for Plan {plan_id}")
            # [FIX] Deduct stock use quantity from inventory
            await process_stock_deduction(db, plan_id=plan_id)
            
            from app.api.utils.mrp import calculate_and_record_mrp
            await calculate_and_record_mrp(db, plan_id=plan_id)
            print(f"MRP Generation Completed for Plan {plan_id}")
        
        # Auto-Complete Logic
        if status == ProductionStatus.COMPLETED:
            affected_po_ids = set()
            affected_oo_ids = set()
            
            for item in plan.items:
                # 0. Update Child Item status to COMPLETED if not already
                if item.status != ProductionStatus.COMPLETED:
                    item.status = ProductionStatus.COMPLETED
                    db.add(item)

                # 1. Update Purchase Items
                for po_item in item.purchase_items:
                    # Set received quantity to full
                    if po_item.quantity > po_item.received_quantity:
                        po_item.received_quantity = po_item.quantity
                        db.add(po_item)
                    affected_po_ids.add(po_item.purchase_order_id)
                    
                # 2. Update Outsourcing Items
                for oo_item in item.outsourcing_items:
                    if oo_item.status != OutsourcingStatus.COMPLETED:
                        oo_item.status = OutsourcingStatus.COMPLETED
                        db.add(oo_item)
                    affected_oo_ids.add(oo_item.outsourcing_order_id)
            
            await db.flush()
            
            # 3. Check and Complete Purchase Orders
            for po_id in affected_po_ids:
                po_query = select(PurchaseOrder).options(selectinload(PurchaseOrder.items)).where(PurchaseOrder.id == po_id)
                po_result = await db.execute(po_query)
                po = po_result.scalars().first()
                
                if po:
                    # Check if all items are fully received
                    all_received = all(i.received_quantity >= i.quantity for i in po.items)
                    if all_received:
                        po.status = PurchaseStatus.COMPLETED
                        po.delivery_date = now_kst().date() # Set actual delivery date?
                        db.add(po)

            # 4. Check and Complete Outsourcing Orders
            for oo_id in affected_oo_ids:
                oo_query = select(OutsourcingOrder).options(selectinload(OutsourcingOrder.items)).where(OutsourcingOrder.id == oo_id)
                oo_result = await db.execute(oo_query)
                oo = oo_result.scalars().first()
                
                if oo:
                    # Check if all items are completed
                    all_completed = all(i.status == OutsourcingStatus.COMPLETED for i in oo.items)
                    if all_completed:
                        oo.status = OutsourcingStatus.COMPLETED
                        oo.delivery_date = now_kst().date()
                        db.add(oo)

            await process_stock_deduction(db, plan_id=plan_id)
            await _handle_production_completion_effects(db, plan)
                
        # Rollback Logic (COMPLETED -> NOT COMPLETED)
        elif old_status == ProductionStatus.COMPLETED and status != ProductionStatus.COMPLETED:
            # 1. Rollback Stocks
            if plan.stock_production:
                sp = plan.stock_production
                # 1-1. 완제품 입고 취소 (차감)
                await handle_stock_movement(
                    db=db,
                    product_id=sp.product_id,
                    quantity=-sp.quantity,
                    transaction_type=TransactionType.OUT,
                    reference=f"Rollback ({sp.production_no})"
                )
                # 1-2. 하위 부품 Backflush 취소 (원복)
                await handle_backflush(
                    db=db,
                    parent_product_id=sp.product_id,
                    produced_quantity=-sp.quantity,
                    reference=f"Rollback ({sp.production_no})"
                )
                # 1-3. 생산 중 수량 복원 (진행 중인 상태로 가는 경우에만 다시 생산 중으로 잡음)
                if status in [ProductionStatus.IN_PROGRESS, ProductionStatus.CONFIRMED]:
                    stock_query = select(Stock).where(Stock.product_id == sp.product_id)
                    s_res = await db.execute(stock_query)
                    stock = s_res.scalars().first()
                    if stock:
                        stock.in_production_quantity += sp.quantity
                
                # Sync StockProduction status
                if status == ProductionStatus.IN_PROGRESS:
                    sp.status = StockProductionStatus.IN_PROGRESS
                elif status == ProductionStatus.CANCELLED:
                    sp.status = StockProductionStatus.CANCELLED
                db.add(sp)
                
            elif plan.order:
                for item in plan.order.items:
                    # 1-1. 완제품 입고 취소 (차감)
                    await handle_stock_movement(
                        db=db,
                        product_id=item.product_id,
                        quantity=-item.quantity,
                        transaction_type=TransactionType.OUT,
                        reference=f"Rollback ({plan.order.order_no})"
                    )
                    # 1-2. 하위 부품 Backflush 취소 (원복)
                    await handle_backflush(
                        db=db,
                        parent_product_id=item.product_id,
                        produced_quantity=-item.quantity,
                        reference=f"Rollback ({plan.order.order_no})"
                    )
                    # 1-3. 생산 중 수량 복원
                    if status in [ProductionStatus.IN_PROGRESS, ProductionStatus.CONFIRMED]:
                        stock_query = select(Stock).where(Stock.product_id == item.product_id)
                        s_res = await db.execute(stock_query)
                        stock = s_res.scalars().first()
                        if stock:
                            stock.in_production_quantity += item.quantity
                
                # Sync Sales Order status
                plan.order.status = OrderStatus.CONFIRMED
                db.add(plan.order)

            # 2. Rollback Linked Orders (Back to PENDING)
            affected_po_ids = set()
            affected_oo_ids = set()
            for item in plan.items:
                for po_item in item.purchase_items:
                    # Revert received quantity? 
                    # (User said "대기 상태로 변경", which usually means received_quantity = 0 or status = PENDING)
                    po_item.received_quantity = 0
                    db.add(po_item)
                    affected_po_ids.add(po_item.purchase_order_id)
                
                for oo_item in item.outsourcing_items:
                    oo_item.status = OutsourcingStatus.PENDING
                    db.add(oo_item)
                    affected_oo_ids.add(oo_item.outsourcing_order_id)
            
            await db.flush()

            # Update PurchaseOrder statuses
            for po_id in affected_po_ids:
                po = await db.get(PurchaseOrder, po_id)
                if po:
                    po.status = PurchaseStatus.PENDING
                    db.add(po)
            
            # Update OutsourcingOrder statuses
            for oo_id in affected_oo_ids:
                oo = await db.get(OutsourcingOrder, oo_id)
                if oo:
                    oo.status = OutsourcingStatus.PENDING
                    db.add(oo)

        elif status == ProductionStatus.IN_PROGRESS:
            if plan.order:
                plan.order.status = OrderStatus.CONFIRMED 
                db.add(plan.order)
            elif plan.stock_production:
                plan.stock_production.status = StockProductionStatus.IN_PROGRESS
                db.add(plan.stock_production)
            
        await db.commit()
        
        # Re-fetch with full options for response
        result = await db.execute(
            select(ProductionPlan)
            .options(
                selectinload(ProductionPlan.items).options(
                    selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                    selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                    selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
                    selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlanItem.work_log_items),  # Fix: MissingGreenlet
                ),
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product),
                selectinload(ProductionPlan.stock_production).selectinload(StockProduction.partner)
            )
            .where(ProductionPlan.id == plan_id)
        )
        return result.scalars().first()
    except MultipleResultsFound:
        raise HTTPException(status_code=400, detail="데이터 중복 오류: 동일 품목에 대해 여러 개의 재고 레코드가 발견되었습니다. 관리자에게 문의하세요.")

@router.patch("/plan-items/{item_id}", response_model=schemas.ProductionPlanItem)
async def update_production_plan_item(
    item_id: int,
    item_in: schemas.ProductionPlanItemUpdate,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Update a single production plan item (status, attachment, etc).
    """
    result = await db.execute(
        select(ProductionPlanItem)
        .options(
            selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
            selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order)
        )
        .where(ProductionPlanItem.id == item_id)
    )
    item = result.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Plan Item not found")

    update_data = item_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(item, field, value)

    # Sync cost for external items if they have linked PO/OO
    if item.course_type in ["PURCHASE", "OUTSOURCING"]:
        total_linked_cost = 0.0
        # Load linked items if not already loaded (though they might be lazy loaded)
        for pi in item.purchase_items:
            total_linked_cost += (pi.quantity * pi.unit_price)
        for oi in item.outsourcing_items:
            total_linked_cost += (oi.quantity * oi.unit_price)
        
        if total_linked_cost > 0:
            item.cost = total_linked_cost

    await db.commit()
    await db.refresh(item)
    
    # --- Auto-Complete Check ---
    if "status" in update_data and update_data["status"] == ProductionStatus.COMPLETED:
        await check_and_complete_production_plan(db, item.plan_id)

        # Link completed external processes to their respective orders
        if item.course_type == "PURCHASE":
            for pi in item.purchase_items:
                if pi.purchase_order and pi.purchase_order.status != PurchaseStatus.COMPLETED:
                    pi.purchase_order.status = PurchaseStatus.COMPLETED
                    pi.purchase_order.actual_delivery_date = now_kst().date()
                    db.add(pi.purchase_order)
            await db.commit()
        elif item.course_type == "OUTSOURCING":
            for oi in item.outsourcing_items:
                if oi.outsourcing_order and oi.outsourcing_order.status != OutsourcingStatus.COMPLETED:
                    oi.outsourcing_order.status = OutsourcingStatus.COMPLETED
                    oi.outsourcing_order.actual_delivery_date = now_kst().date()
                    db.add(oi.outsourcing_order)
            await db.commit()
    
    # Reload for full schema
    result = await db.execute(
        select(ProductionPlanItem)
        .options(
            selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
            selectinload(ProductionPlanItem.plan).options(
                selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                selectinload(ProductionPlan.stock_production).options(
                    selectinload(StockProduction.product).options(
                        selectinload(Product.standard_processes).selectinload(ProductProcess.process)
                    ),
                    selectinload(StockProduction.partner)
                )
            ),
            selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
            selectinload(ProductionPlanItem.work_log_items)
        )
        .where(ProductionPlanItem.id == item_id)
    )
    item = result.scalars().first()
    item.completed_quantity = calculate_completed_quantity(item)
    return item

# --- Work Log Endpoints ---

@router.get("/work-logs", response_model=List[schemas.WorkLog])
async def read_work_logs(
    skip: int = 0,
    limit: int = 1000,
    major_group_id: Optional[int] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: Staff = Depends(deps.get_current_user)
) -> Any:
    """
    Retrieve work logs.
    """
    stmt = select(WorkLog)
    
    # 일반 사용자의 경우 본인이 작성자(worker_id)인 일지만 조회 가능
    if current_user.user_type != "ADMIN":
        stmt = stmt.where(WorkLog.worker_id == current_user.id)
        
    if major_group_id:
        subquery = select(WorkLogItem.work_log_id).join(ProductionPlanItem).join(Product).join(ProductGroup, Product.group_id == ProductGroup.id)\
                     .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))
        stmt = stmt.where(WorkLog.id.in_(subquery))

    result = await db.execute(
        stmt
        .options(
            selectinload(WorkLog.worker),
            selectinload(WorkLog.items).options(
                selectinload(WorkLogItem.worker),
                selectinload(WorkLogItem.work_log),
                selectinload(WorkLogItem.plan_item).options(
                    selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                    selectinload(ProductionPlanItem.equipment),
                    selectinload(ProductionPlanItem.worker),
                    selectinload(ProductionPlanItem.purchase_items),
                    selectinload(ProductionPlanItem.outsourcing_items),
                    selectinload(ProductionPlanItem.plan).options(
                        selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                        selectinload(ProductionPlan.stock_production).options(
                            selectinload(StockProduction.product),
                            selectinload(StockProduction.partner)
                        )
                    )
                )
            )
        )
        .order_by(WorkLog.work_date.desc())
        .offset(skip).limit(limit)
    )
    return result.scalars().all()

@router.post("/work-logs", response_model=schemas.WorkLog)
async def create_work_log(
    log_in: schemas.WorkLogCreate,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Create a new work log.
    """
    import json
    from fastapi import HTTPException
    
    # Check for existing log on the same date for the same worker
    stmt = select(WorkLog).where(
        WorkLog.work_date == log_in.work_date,
        WorkLog.worker_id == log_in.worker_id
    )
    result = await db.execute(stmt)
    existing_log = result.scalars().first()

    if existing_log and log_in.mode == "CREATE":
        raise HTTPException(status_code=409, detail="해당 날짜에 이미 등록된 작업일지가 있습니다.")

    if existing_log and log_in.mode == "REPLACE":
        await db.delete(existing_log)
        await db.flush()
        existing_log = None

    if existing_log and log_in.mode == "MERGE":
        log = existing_log
        if log_in.note:
            log.note = (log.note or "") + "\n" + log_in.note
        # Handle attachments (merge lists)
        if log_in.attachment_file:
            current_files = []
            if log.attachment_file:
                if isinstance(log.attachment_file, str):
                    current_files = json.loads(log.attachment_file)
                else:
                    current_files = log.attachment_file
            
            new_files = log_in.attachment_file
            if isinstance(new_files, str):
                new_files = json.loads(new_files)
                
            combined = current_files + new_files
            log.attachment_file = json.dumps(combined, ensure_ascii=False)
    else:
        attachment_file = log_in.attachment_file
        if attachment_file and isinstance(attachment_file, (list, dict)):
            attachment_file = json.dumps(attachment_file, ensure_ascii=False)

        log = WorkLog(
            work_date=log_in.work_date,
            worker_id=log_in.worker_id,
            note=log_in.note,
            attachment_file=attachment_file
        )
        db.add(log)
        await db.flush()

    for item_in in log_in.items:
        # Fetch plan item to get default price if needed
        plan_item = await db.get(ProductionPlanItem, item_in.plan_item_id)
        u_price = item_in.unit_price
        if not u_price and plan_item:
            u_price = (plan_item.cost or 0) / (plan_item.quantity or 1)

        log_item = WorkLogItem(
            work_log_id=log.id,
            plan_item_id=item_in.plan_item_id,
            worker_id=item_in.worker_id,
            start_time=item_in.start_time,
            end_time=item_in.end_time,
            good_quantity=item_in.good_quantity,
            bad_quantity=item_in.bad_quantity,
            unit_price=u_price,
            note=item_in.note
        )
        db.add(log_item)
        await db.flush()
        await sync_plan_item_status(db, item_in.plan_item_id)
        
        # --- Stock Movement & Backflush Hook ---
        if item_in.good_quantity > 0 and plan_item:
            # 1. 완제품/공정품 입고 처리
            await handle_stock_movement(
                db=db,
                product_id=plan_item.product_id,
                quantity=item_in.good_quantity,
                transaction_type=TransactionType.IN,
                reference=f"WorkLog (Item {log_item.id})"
            )
            # 2. 하위 부품 Backflush 처리
            await handle_backflush(
                db=db,
                parent_product_id=plan_item.product_id,
                produced_quantity=item_in.good_quantity,
                reference=f"WorkLog (Item {log_item.id})"
            )

    await db.commit()
    await db.refresh(log)

    result = await db.execute(
        select(WorkLog)
        .options(
            selectinload(WorkLog.worker),
            selectinload(WorkLog.items).options(
                selectinload(WorkLogItem.worker),
                selectinload(WorkLogItem.work_log),
                selectinload(WorkLogItem.plan_item).options(
                    selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                    selectinload(ProductionPlanItem.equipment),
                    selectinload(ProductionPlanItem.worker),
                    selectinload(ProductionPlanItem.purchase_items),
                    selectinload(ProductionPlanItem.outsourcing_items),
                    selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
                )
            )
        )
        .where(WorkLog.id == log.id)
    )
    return result.scalars().first()

@router.put("/work-logs/{log_id}", response_model=schemas.WorkLog)
async def update_work_log(
    log_id: int,
    log_in: schemas.WorkLogUpdate,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Update a work log.
    """
    result = await db.execute(
        select(WorkLog).options(selectinload(WorkLog.items)).where(WorkLog.id == log_id)
    )
    log = result.scalars().first()
    if not log:
        raise HTTPException(status_code=404, detail="Work Log not found")

    if log_in.work_date is not None:
        log.work_date = log_in.work_date
    if log_in.worker_id is not None:
        log.worker_id = log_in.worker_id
    if log_in.note is not None:
        log.note = log_in.note
    if log_in.attachment_file is not None:
        import json
        if isinstance(log_in.attachment_file, (list, dict)):
            log.attachment_file = json.dumps(log_in.attachment_file, ensure_ascii=False)
        else:
            log.attachment_file = log_in.attachment_file

    if log_in.items is not None:
        # --- Stock Reversal Hook ---
        for old_item in log.items:
            if old_item.good_quantity > 0:
                # Load plan item to get product_id
                plan_item = await db.get(ProductionPlanItem, old_item.plan_item_id)
                if plan_item:
                    # 완제품 입고 취소 (출고 처리)
                    await handle_stock_movement(
                        db=db,
                        product_id=plan_item.product_id,
                        quantity=-old_item.good_quantity,
                        transaction_type=TransactionType.ADJUSTMENT,
                        reference=f"WorkLog Update (Reverse Item {old_item.id})"
                    )
                    # 하위 부품 Backflush 취소 (입고 처리)
                    await handle_backflush(
                        db=db,
                        parent_product_id=plan_item.product_id,
                        produced_quantity=-old_item.good_quantity,
                        reference=f"WorkLog Update (Reverse Item {old_item.id})"
                    )

        log.items.clear()
        
        for item_in in log_in.items:
            # Fetch plan item to get default price if needed
            plan_item = await db.get(ProductionPlanItem, item_in.plan_item_id)
            u_price = item_in.unit_price
            if not u_price and plan_item:
                u_price = (plan_item.cost or 0) / (plan_item.quantity or 1)

            log_item = WorkLogItem(
                work_log_id=log.id,
                plan_item_id=item_in.plan_item_id,
                worker_id=item_in.worker_id,
                start_time=item_in.start_time,
                end_time=item_in.end_time,
                good_quantity=item_in.good_quantity,
                bad_quantity=item_in.bad_quantity,
                unit_price=u_price,
                note=item_in.note
            )
            log.items.append(log_item)
            await db.flush()
            await sync_plan_item_status(db, item_in.plan_item_id)

            # --- Stock Movement & Backflush Hook for NEW items ---
            if item_in.good_quantity > 0 and plan_item:
                await handle_stock_movement(
                    db=db,
                    product_id=plan_item.product_id,
                    quantity=item_in.good_quantity,
                    transaction_type=TransactionType.IN,
                    reference=f"WorkLog Update (New Item {log_item.id})"
                )
                await handle_backflush(
                    db=db,
                    parent_product_id=plan_item.product_id,
                    produced_quantity=item_in.good_quantity,
                    reference=f"WorkLog Update (New Item {log_item.id})"
                )

    await db.commit()

    result = await db.execute(
        select(WorkLog)
        .options(
            selectinload(WorkLog.worker),
            selectinload(WorkLog.items).options(
                selectinload(WorkLogItem.worker),
                selectinload(WorkLogItem.work_log),
                selectinload(WorkLogItem.plan_item).options(
                    selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                    selectinload(ProductionPlanItem.equipment),
                    selectinload(ProductionPlanItem.worker),
                    selectinload(ProductionPlanItem.purchase_items),
                    selectinload(ProductionPlanItem.outsourcing_items),
                    selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlanItem.plan).selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
                )
            )
        )
        .where(WorkLog.id == log.id)
    )
    return result.scalars().first()

@router.delete("/work-logs/{log_id}")
async def delete_work_log(
    log_id: int,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Delete a work log.
    """
    result = await db.execute(select(WorkLog).where(WorkLog.id == log_id))
    log = result.scalars().first()
    if not log:
        raise HTTPException(status_code=404, detail="Work Log not found")

    # Store item IDs to sync status after deletion
    plan_item_ids = [item.plan_item_id for item in log.items if item.plan_item_id]

    # --- Stock Reversal Hook ---
    for item in log.items:
        if item.good_quantity > 0:
            plan_item = await db.get(ProductionPlanItem, item.plan_item_id)
            if plan_item:
                await handle_stock_movement(
                    db=db,
                    product_id=plan_item.product_id,
                    quantity=-item.good_quantity,
                    transaction_type=TransactionType.ADJUSTMENT,
                    reference=f"WorkLog Delete (Reverse Item {item.id})"
                )
                await handle_backflush(
                    db=db,
                    parent_product_id=plan_item.product_id,
                    produced_quantity=-item.good_quantity,
                    reference=f"WorkLog Delete (Reverse Item {item.id})"
                )

    await db.delete(log)
    await db.commit()
    
    # Sync status for all affected items
    for pid in plan_item_ids:
        await sync_plan_item_status(db, pid)
    await db.commit()

    return {"message": "Work Log deleted successfully"}

# --- Performance Management Endpoints ---

@router.get("/performance/workers")
async def get_worker_performance(
    start_date: Union[date, None] = None,
    end_date: Union[date, None] = None,
    worker_id: Optional[int] = None,
    major_group_id: Optional[int] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: Staff = Depends(deps.get_current_user)
) -> Any:
    """
    Get aggregated performance by worker with optional date and worker filtering.
    """
    from sqlalchemy import func, distinct, case
    
    # Cost with fallback: if WorkLogItem.unit_price is 0 or null, try (PlanItem.cost / PlanItem.quantity)
    # Using CASE to handle potential division by zero
    calc_unit_price = case(
        (WorkLogItem.unit_price > 0, WorkLogItem.unit_price),
        (ProductionPlanItem.quantity > 0, ProductionPlanItem.cost / ProductionPlanItem.quantity),
        else_=0
    )
    
    stmt = (
        select(
            Staff.id.label("worker_id"),
            Staff.name.label("worker_name"),
            func.sum(WorkLogItem.good_quantity * calc_unit_price).label("total_cost"),
            func.count(distinct(WorkLog.work_date)).label("log_days")
        )
        .join(WorkLogItem, Staff.id == WorkLogItem.worker_id)
        .join(WorkLog, WorkLogItem.work_log_id == WorkLog.id)
        .join(ProductionPlanItem, WorkLogItem.plan_item_id == ProductionPlanItem.id)
        .group_by(Staff.id, Staff.name)
    )
    
    if start_date:
        stmt = stmt.where(WorkLog.work_date >= start_date)
    if end_date:
        stmt = stmt.where(WorkLog.work_date <= end_date)
    if worker_id:
        stmt = stmt.where(Staff.id == worker_id)
        
    if major_group_id:
        from app.models.product import ProductGroup
        from sqlalchemy import or_
        # Note: ProductionPlanItem and Product are already joined for calc_unit_price/total_cost
        stmt = stmt.join(ProductGroup, Product.group_id == ProductGroup.id)\
                   .where(or_(ProductGroup.id == major_group_id, ProductGroup.parent_id == major_group_id))
        
    # 일반 사용자의 경우 본인 데이터만 조회
    if current_user.user_type != "ADMIN":
        stmt = stmt.where(Staff.id == current_user.id)
        
    result = await db.execute(stmt)
    return [dict(row._mapping) for row in result.all()]

@router.get("/performance/details", response_model=List[schemas.WorkLogItem])
async def get_performance_details(
    worker_id: Optional[int] = None,
    start_date: Union[date, None] = None,
    end_date: Union[date, None] = None,
    db: AsyncSession = Depends(deps.get_db),
    current_user: Staff = Depends(deps.get_current_user)
) -> Any:
    """
    Get detailed work log items with optional worker and date filtering.
    """
    stmt = (
        select(WorkLogItem)
        .join(WorkLog, WorkLogItem.work_log_id == WorkLog.id)
        .options(
            selectinload(WorkLogItem.work_log),
            selectinload(WorkLogItem.worker),
            selectinload(WorkLogItem.plan_item).options(
                selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(ProductionPlanItem.equipment),
                selectinload(ProductionPlanItem.worker),
                selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
                selectinload(ProductionPlanItem.plan).options(
                    selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
                )
            )
        )
    )
    
    if worker_id:
        stmt = stmt.where(WorkLogItem.worker_id == worker_id)
    if start_date:
        stmt = stmt.where(WorkLog.work_date >= start_date)
    if end_date:
        stmt = stmt.where(WorkLog.work_date <= end_date)
        
    # 일반 사용자의 경우 본인 데이터만 조회
    if current_user.user_type != "ADMIN":
        stmt = stmt.where(WorkLogItem.worker_id == current_user.id)
        
    result = await db.execute(stmt.order_by(WorkLog.work_date.desc(), WorkLogItem.id.desc()))
    return result.scalars().all()

@router.get("/performance/workers/{worker_id}/details", response_model=List[schemas.WorkLogItem])
async def get_worker_performance_details(
    worker_id: int,
    start_date: Union[date, None] = None,
    end_date: Union[date, None] = None,
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Get detailed work log items for a specific worker with optional date filtering.
    """
    stmt = (
        select(WorkLogItem)
        .join(WorkLog, WorkLogItem.work_log_id == WorkLog.id)
        .options(
            selectinload(WorkLogItem.work_log),
            selectinload(WorkLogItem.worker),
            selectinload(WorkLogItem.plan_item).options(
                selectinload(ProductionPlanItem.product).selectinload(Product.standard_processes).selectinload(ProductProcess.process),
                selectinload(ProductionPlanItem.equipment),
                selectinload(ProductionPlanItem.worker),
                selectinload(ProductionPlanItem.purchase_items).selectinload(PurchaseOrderItem.purchase_order),
                selectinload(ProductionPlanItem.outsourcing_items).selectinload(OutsourcingOrderItem.outsourcing_order),
                selectinload(ProductionPlanItem.plan).options(
                    selectinload(ProductionPlan.order).selectinload(SalesOrder.partner),
                    selectinload(ProductionPlan.stock_production).selectinload(StockProduction.product)
                )
            )
        )
        .where(WorkLogItem.worker_id == worker_id)
    )
    
    if start_date:
        stmt = stmt.where(WorkLog.work_date >= start_date)
    if end_date:
        stmt = stmt.where(WorkLog.work_date <= end_date)
        
    result = await db.execute(stmt.order_by(WorkLog.work_date.desc(), WorkLogItem.id.desc()))
    return result.scalars().all()

@router.patch("/work-log-items/{item_id}", response_model=schemas.WorkLogItem)
async def update_work_log_item(
    item_id: int,
    item_in: schemas.WorkLogItemBase, # Using Base as it contains all editable fields
    db: AsyncSession = Depends(deps.get_db),
) -> Any:
    """
    Update a single work log item (price, quantity, worker, etc).
    """
    result = await db.execute(
        select(WorkLogItem)
        .options(selectinload(WorkLogItem.work_log))
        .where(WorkLogItem.id == item_id)
    )
    item = result.scalars().first()
    if not item:
        raise HTTPException(status_code=404, detail="Work Log Item not found")

    update_data = item_in.model_dump(exclude_unset=True)
    
    # Identify if quantity or price changed to sync plan item status/cost
    qty_changed = "good_quantity" in update_data and update_data["good_quantity"] != item.good_quantity
    
    old_good_qty = item.good_quantity
    
    for field, value in update_data.items():
        setattr(item, field, value)

    await db.commit()
    await db.refresh(item)
    
    if qty_changed:
        await sync_plan_item_status(db, item.plan_item_id)
        
        # --- Stock Adjustment Hook for DIFF ---
        diff_qty = item.good_quantity - old_good_qty
        if diff_qty != 0:
            plan_item = await db.get(ProductionPlanItem, item.plan_item_id)
            if plan_item:
                await handle_stock_movement(
                    db=db,
                    product_id=plan_item.product_id,
                    quantity=diff_qty,
                    transaction_type=TransactionType.ADJUSTMENT,
                    reference=f"WorkLogItem Update (Item {item.id})"
                )
                await handle_backflush(
                    db=db,
                    parent_product_id=plan_item.product_id,
                    produced_quantity=diff_qty,
                    reference=f"WorkLogItem Update (Item {item.id})"
                )
        await db.commit()

    # Re-fetch for full schema
    result = await db.execute(
        select(WorkLogItem)
        .options(
            selectinload(WorkLogItem.work_log),
            selectinload(WorkLogItem.plan_item).options(
                selectinload(ProductionPlanItem.product),
                selectinload(ProductionPlanItem.plan).options(
                    selectinload(ProductionPlan.order),
                    selectinload(ProductionPlan.stock_production)
                )
            )
        )
        .where(WorkLogItem.id == item_id)
    )
    return result.scalars().first()
